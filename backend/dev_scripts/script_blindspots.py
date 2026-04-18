"""
Test the blind spots — scenarios most likely to break in production.

1. Multi-document SOV (combine multiple PDFs, single extraction)
2. Baseline update with conflicting/overlapping data
3. Vague/bad field descriptions (what real users type)
4. Documents where columns don't match field names at all
5. Very small documents (1-page SOV with 2-3 rows)
6. Papyra renewal quote PDFs (real-world broker docs)
7. Mixed document types in same extraction
8. Fields with numeric flag but doc has text values
9. Re-extraction: same doc, different field sets, verify no cross-contamination
10. The actual pipeline poller path (auto routing, no force_sov)
"""
import asyncio, sys, os, time, json, traceback
sys.path.insert(0, os.path.dirname(__file__))
sys.stdout.reconfigure(line_buffering=True)

import logging
logging.basicConfig(level=logging.WARNING, format="%(levelname)s %(name)s: %(message)s")
logging.getLogger("httpx").setLevel(logging.ERROR)
logging.getLogger("httpcore").setLevel(logging.ERROR)
logging.getLogger("openai").setLevel(logging.ERROR)

from app.config import settings
settings.mistral_api_key = ""

from app.services.pdf_service import parse_pdf, combine_parsed_documents
from app.services.extraction import extract_from_document, LLMUsage
from app.services.spreadsheet_service import (
    generate_excel_bytes, generate_csv_bytes, read_headers_from_bytes,
    update_excel_baseline_bytes, update_csv_baseline_bytes,
)

TEST_DOCS = "/Users/martinmashalov/Downloads/GridPull/test_documents"
TEST_FILES = "/Users/martinmashalov/Downloads/GridPull/backend/test_files"
PAPYRA = "/Users/martinmashalov/Downloads/Papyra/fixtures/renewal_quote_test"
EMPTY = {"", "null", "none", "n/a", "na", "-", "—", "unknown", "not found", "not available"}
results = []

def is_filled(val):
    if val is None: return False
    return str(val).strip().lower() not in EMPTY

def fill_stats(rows, fn):
    total = len(rows) * len(fn)
    filled = sum(1 for r in rows for f in fn if is_filled(r.get(f)))
    return filled, total, (filled/total*100) if total else 0

async def run(tid, label, path_or_paths, fields, *, instructions="", force_sov=True,
              force_general=False, min_rows=1, multi_doc=False, batch_count=1):
    fn = [f["name"] for f in fields]
    print(f"\n[T{tid:02d}] {label}")
    t0 = time.time()
    try:
        if multi_doc and isinstance(path_or_paths, list):
            docs = []
            for p in path_or_paths:
                d = await asyncio.to_thread(parse_pdf, p, os.path.basename(p))
                docs.append(d)
                print(f"  Parsed: {os.path.basename(p)} ({d.page_count} pages)")
            combined = combine_parsed_documents(docs)
            print(f"  Combined: {combined.page_count} pages, {len(combined.tables)} tables")
            usage = LLMUsage()
            rows = await extract_from_document(combined, fields, usage, instructions,
                                               force_sov=force_sov, force_general=force_general,
                                               batch_document_count=1)
        else:
            path = path_or_paths if isinstance(path_or_paths, str) else path_or_paths[0]
            doc = await asyncio.to_thread(parse_pdf, path, os.path.basename(path))
            print(f"  Parsed: {doc.page_count} pages, type={doc.doc_type_hint}")
            usage = LLMUsage()
            rows = await extract_from_document(doc, fields, usage, instructions,
                                               force_sov=force_sov, force_general=force_general,
                                               batch_document_count=batch_count)

        elapsed = time.time() - t0
        filled, total, pct = fill_stats(rows, fn)

        # Validate spreadsheet
        xlsx_ok = csv_ok = True
        try:
            generate_excel_bytes(rows, fn)
        except Exception as e:
            xlsx_ok = False
        try:
            generate_csv_bytes(rows, fn)
        except Exception as e:
            csv_ok = False

        status = "PASS" if len(rows) >= min_rows and xlsx_ok and csv_ok else "FAIL"
        print(f"  {status}: {len(rows)} rows, {pct:.0f}% fill, ${usage.cost_usd:.4f}, {elapsed:.1f}s")
        if rows:
            sample = {k: str(v)[:35] for k,v in rows[0].items() if k not in ("_source_file","_error") and is_filled(v)}
            print(f"  Row1: {json.dumps(sample, default=str)[:200]}")
        if len(rows) > 1:
            sample2 = {k: str(v)[:35] for k,v in rows[-1].items() if k not in ("_source_file","_error") and is_filled(v)}
            print(f"  Last: {json.dumps(sample2, default=str)[:200]}")
        results.append({"id": tid, "label": label, "status": status, "rows": len(rows), "fill": round(pct,1)})
        return rows
    except Exception as e:
        elapsed = time.time() - t0
        print(f"  CRASH: {e}")
        traceback.print_exc()
        results.append({"id": tid, "label": label, "status": "CRASH", "rows": 0, "fill": 0})
        return []


async def main():
    tid = 0

    print("=" * 80)
    print("BLIND SPOT TESTS")
    print("=" * 80)

    SOV_FIELDS = [
        {"name": "Location #", "description": "Location number"},
        {"name": "Address", "description": "Street address"},
        {"name": "City", "description": "City"},
        {"name": "State", "description": "State"},
        {"name": "Zip Code", "description": "ZIP code"},
        {"name": "Building Value", "description": "Building value", "numeric": True},
        {"name": "Total Insured Value", "description": "TIV", "numeric": True},
    ]

    # ── 1. Multi-doc SOV: combine 2 PDFs into one extraction ─────────────
    tid += 1
    await run(tid, "Multi-doc SOV: 25-bldg + 50-loc combined", [
        f"{TEST_FILES}/01_property_appraisal_report_25_buildings.pdf",
        f"{TEST_FILES}/02_carrier_property_schedule_50_locations.pdf",
    ], SOV_FIELDS, multi_doc=True, min_rows=50)

    # ── 2. Multi-doc SOV: property + vehicle (different schemas) ─────────
    tid += 1
    await run(tid, "Multi-doc: property + vehicle (schema mismatch)", [
        f"{TEST_FILES}/01_property_appraisal_report_25_buildings.pdf",
        f"{TEST_FILES}/04_vehicle_schedule_35_vehicles.pdf",
    ], SOV_FIELDS, multi_doc=True, min_rows=20)

    # ── 3. Vague/lazy field descriptions (what users actually type) ──────
    tid += 1
    await run(tid, "Vague descriptions (lazy user)", f"{TEST_FILES}/01_property_appraisal_report_25_buildings.pdf", [
        {"name": "Location #", "description": "location"},
        {"name": "Address", "description": "address"},
        {"name": "City", "description": "city"},
        {"name": "State", "description": "state"},
        {"name": "Value", "description": "the value"},
    ], min_rows=20)

    # ── 4. Field names that are totally wrong for the document ───────────
    tid += 1
    await run(tid, "Completely wrong fields for doc", f"{TEST_FILES}/01_property_appraisal_report_25_buildings.pdf", [
        {"name": "Patient Name", "description": "Patient name"},
        {"name": "Diagnosis Code", "description": "ICD-10 code"},
        {"name": "Copay Amount", "description": "Patient copay", "numeric": True},
        {"name": "Provider NPI", "description": "National provider identifier"},
    ], min_rows=1)

    # ── 5. SOV with auto routing (no force_sov — pipeline poller path) ──
    tid += 1
    await run(tid, "Auto-route: SOV fields on SOV doc", f"{TEST_FILES}/02_carrier_property_schedule_50_locations.pdf",
        SOV_FIELDS, force_sov=False, force_general=False, min_rows=40)

    # ── 6. Auto-route: invoice fields on invoice (should NOT go to SOV) ──
    tid += 1
    invoice_fields = [
        {"name": "Invoice Number", "description": "Invoice #"},
        {"name": "Date", "description": "Invoice date"},
        {"name": "Customer Name", "description": "Customer"},
        {"name": "Total Amount", "description": "Total due", "numeric": True},
    ]
    await run(tid, "Auto-route: invoice fields on invoice", f"{TEST_DOCS}/01_invoices/invoice_Aaron Bergman_36258.pdf",
        invoice_fields, force_sov=False, force_general=False, batch_count=3, min_rows=1)

    # ── 7. Baseline update: overlapping locations (merge test) ───────────
    tid += 1
    print(f"\n[T{tid:02d}] Baseline merge: overlapping locations")
    try:
        fn = [f["name"] for f in SOV_FIELDS]
        # Existing baseline with 3 locations
        existing = [
            {"_source_file": "old.pdf", "Location #": "1", "Address": "100 Main St", "City": "Dallas", "State": "TX", "Zip Code": "75201", "Building Value": "1000000", "Total Insured Value": "1500000"},
            {"_source_file": "old.pdf", "Location #": "2", "Address": "200 Oak Ave", "City": "Houston", "State": "TX", "Zip Code": "77001", "Building Value": "2000000", "Total Insured Value": "3000000"},
            {"_source_file": "old.pdf", "Location #": "3", "Address": "300 Pine Rd", "City": "Austin", "State": "TX", "Zip Code": "78701", "Building Value": "500000", "Total Insured Value": "750000"},
        ]
        baseline = generate_excel_bytes(existing, fn)

        # New extraction with overlapping + new locations
        new_rows = [
            {"_source_file": "new.pdf", "Location #": "2", "Address": "200 Oak Ave", "City": "Houston", "State": "TX", "Zip Code": "77001", "Building Value": "2200000", "Total Insured Value": "3300000"},
            {"_source_file": "new.pdf", "Location #": "4", "Address": "400 Elm St", "City": "Denver", "State": "CO", "Zip Code": "80201", "Building Value": "800000", "Total Insured Value": "1200000"},
        ]

        # Test with allow_edit=False (preserve existing values)
        updated = update_excel_baseline_bytes(baseline, new_rows, fn, allow_edit_past_values=False)
        hdrs = read_headers_from_bytes(updated, "xlsx")
        import openpyxl, io
        wb = openpyxl.load_workbook(io.BytesIO(updated))
        ws = wb.active
        row_count = ws.max_row - 1  # minus header
        # Check Loc 2 wasn't overwritten
        loc2_bv = None
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(r, 1).value).strip() in ("2", "old.pdf"):
                # Find Building Value column
                for c in range(1, ws.max_column + 1):
                    if ws.cell(1, c).value == "Building Value":
                        loc2_bv = ws.cell(r, c).value
                        break
        print(f"  Rows: {row_count} | Headers: {hdrs}")
        print(f"  Loc 2 BV (should be 2000000, not overwritten): {loc2_bv}")
        ok = row_count >= 4 and "GridPull Status" in hdrs
        print(f"  {'PASS' if ok else 'FAIL'}")

        # Test with allow_edit=True (overwrite existing)
        updated2 = update_excel_baseline_bytes(baseline, new_rows, fn, allow_edit_past_values=True)
        wb2 = openpyxl.load_workbook(io.BytesIO(updated2))
        ws2 = wb2.active
        loc2_bv2 = None
        for r in range(2, ws2.max_row + 1):
            for c in range(1, ws2.max_column + 1):
                if ws2.cell(1, c).value == "Location #" and str(ws2.cell(r, c).value).strip() == "2":
                    for c2 in range(1, ws2.max_column + 1):
                        if ws2.cell(1, c2).value == "Building Value":
                            loc2_bv2 = ws2.cell(r, c2).value
                            break
        print(f"  Edit mode Loc 2 BV (should be 2200000, overwritten): {loc2_bv2}")

        results.append({"id": tid, "label": "Baseline overlap merge", "status": "PASS" if ok else "FAIL", "rows": row_count, "fill": 0})
    except Exception as e:
        print(f"  CRASH: {e}")
        traceback.print_exc()
        results.append({"id": tid, "label": "Baseline overlap merge", "status": "CRASH", "rows": 0, "fill": 0})

    # ── 8. Numeric flag but text values in doc ───────────────────────────
    tid += 1
    await run(tid, "Numeric fields on text-heavy doc", f"{TEST_DOCS}/09_contracts_agreements/contract_001.pdf", [
        {"name": "Contract Value", "description": "Total contract value", "numeric": True},
        {"name": "Duration Years", "description": "Contract duration in years", "numeric": True},
        {"name": "Party Count", "description": "Number of parties", "numeric": True},
        {"name": "Contract Title", "description": "Title of the agreement"},
    ], force_sov=False, force_general=True, min_rows=1)

    # ── 9. Cross-contamination: same doc, 2 different field sets ─────────
    tid += 1
    print(f"\n[T{tid:02d}] Cross-contamination check")
    doc = await asyncio.to_thread(parse_pdf, f"{TEST_FILES}/01_property_appraisal_report_25_buildings.pdf",
                                   "01_property_appraisal_report_25_buildings.pdf")
    usage1 = LLMUsage()
    rows1 = await extract_from_document(doc, [
        {"name": "Address", "description": "Street address"},
        {"name": "City", "description": "City"},
    ], usage1, force_sov=True)

    usage2 = LLMUsage()
    rows2 = await extract_from_document(doc, [
        {"name": "Building Value", "description": "Building value", "numeric": True},
        {"name": "Year Built", "description": "Year constructed"},
    ], usage2, force_sov=True)

    # Check no Address/City leaked into rows2 and no Value/Year into rows1
    leaked1 = any(r.get("Building Value") or r.get("Year Built") for r in rows1)
    leaked2 = any(r.get("Address") or r.get("City") for r in rows2)
    ok = not leaked1 and not leaked2 and len(rows1) >= 20 and len(rows2) >= 20
    print(f"  Set1 (Addr/City): {len(rows1)} rows, leaked_values={leaked1}")
    print(f"  Set2 (Value/Year): {len(rows2)} rows, leaked_values={leaked2}")
    print(f"  {'PASS' if ok else 'FAIL'}: no cross-contamination")
    results.append({"id": tid, "label": "Cross-contamination", "status": "PASS" if ok else "FAIL", "rows": len(rows1)+len(rows2), "fill": 0})

    # ── 10. Papyra renewal quote PDFs (real broker docs) ─────────────────
    if os.path.isdir(PAPYRA):
        for fname in sorted(os.listdir(PAPYRA)):
            if not fname.endswith(".pdf"):
                continue
            tid += 1
            await run(tid, f"Papyra: {fname}",
                os.path.join(PAPYRA, fname), [
                    {"name": "Insured Name", "description": "Name of the insured"},
                    {"name": "Policy Number", "description": "Policy number"},
                    {"name": "Effective Date", "description": "Policy effective date"},
                    {"name": "Expiration Date", "description": "Policy expiration date"},
                    {"name": "Premium", "description": "Total premium amount", "numeric": True},
                    {"name": "Coverage Type", "description": "Type of coverage"},
                    {"name": "Limit", "description": "Coverage limit", "numeric": True},
                    {"name": "Deductible", "description": "Deductible amount", "numeric": True},
                ], force_sov=True, min_rows=1)

    # ── 11. SOV with conflicting instructions ────────────────────────────
    tid += 1
    await run(tid, "Conflicting instructions", f"{TEST_FILES}/01_property_appraisal_report_25_buildings.pdf",
        SOV_FIELDS, instructions="Only extract the first 5 locations. Ignore all locations after #5.", min_rows=1)

    # ── 12. Very small PDF (package policy - 15 locations, dec pages) ────
    tid += 1
    await run(tid, "Dec pages 15 locations", f"{TEST_FILES}/03_package_policy_dec_pages_15_locations.pdf", [
        {"name": "Location #", "description": "Location number"},
        {"name": "Address", "description": "Address"},
        {"name": "City", "description": "City"},
        {"name": "State", "description": "State"},
        {"name": "Zip Code", "description": "Zip"},
        {"name": "Building Value", "description": "Building value", "numeric": True},
        {"name": "TIV", "description": "Total insured value", "numeric": True},
        {"name": "Occupancy", "description": "Occupancy type"},
        {"name": "Construction", "description": "Construction type"},
        {"name": "Year Built", "description": "Year built"},
        {"name": "Sprinklered", "description": "Has sprinklers Y/N"},
        {"name": "Stories", "description": "Number of stories"},
    ], min_rows=10)

    # ── 13. Prior year SOV with full field set ───────────────────────────
    tid += 1
    await run(tid, "Prior year SOV 30 locations (full fields)", f"{TEST_FILES}/05_prior_year_sov_30_locations.pdf", [
        {"name": "Location #", "description": "Location number"},
        {"name": "Address", "description": "Street address"},
        {"name": "City", "description": "City"},
        {"name": "State", "description": "State"},
        {"name": "Zip Code", "description": "Zip code"},
        {"name": "Building Value", "description": "Building replacement cost", "numeric": True},
        {"name": "Contents Value", "description": "Contents/BPP value", "numeric": True},
        {"name": "Business Income", "description": "Business income value", "numeric": True},
        {"name": "Total Insured Value", "description": "TIV", "numeric": True},
        {"name": "Year Built", "description": "Year constructed"},
        {"name": "Square Footage", "description": "Total square footage", "numeric": True},
        {"name": "Construction Type", "description": "Construction class"},
        {"name": "Occupancy", "description": "Occupancy/use type"},
        {"name": "Sprinklered", "description": "Sprinkler protection"},
        {"name": "Stories", "description": "Number of floors"},
        {"name": "Roof Type", "description": "Roof material"},
    ], min_rows=25)

    # ── 14. 20-location full SOV ─────────────────────────────────────────
    f = f"{TEST_FILES}/06_full_sov_20_locations.pdf"
    if os.path.exists(f):
        tid += 1
        await run(tid, "Full SOV 20 locations", f, SOV_FIELDS, min_rows=15)

    # ── 15. Customer intake form (not a traditional SOV) ─────────────────
    f = f"{TEST_FILES}/07_customer_intake_form.pdf"
    if os.path.exists(f):
        tid += 1
        await run(tid, "Customer intake form", f, SOV_FIELDS, min_rows=1)

    # ── 16. Appraisal supplement ─────────────────────────────────────────
    f = f"{TEST_FILES}/08_appraisal_supplement.pdf"
    if os.path.exists(f):
        tid += 1
        await run(tid, "Appraisal supplement", f, SOV_FIELDS, min_rows=1)

    # ── REPORT ───────────────────────────────────────────────────────────
    print(f"\n{'='*80}")
    print("BLIND SPOT TEST REPORT")
    print(f"{'='*80}")
    passed = sum(1 for r in results if r["status"] == "PASS")
    failed = sum(1 for r in results if r["status"] == "FAIL")
    crashed = sum(1 for r in results if r["status"] == "CRASH")
    print(f"\nTotal: {len(results)} | PASS: {passed} | FAIL: {failed} | CRASH: {crashed}")
    for r in results:
        s = {"PASS": "  ok", "FAIL": "FAIL", "CRASH": "!ERR"}[r["status"]]
        print(f"  T{r['id']:02d} {s} rows={r['rows']:>3} fill={r['fill']:>5.0f}% {r['label']}")
    if crashed:
        print(f"\n*** {crashed} CRASHES — FIX REQUIRED ***")
    print(f"{'='*80}")


if __name__ == "__main__":
    asyncio.run(main())
