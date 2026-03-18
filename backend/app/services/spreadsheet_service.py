import io
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import csv
from datetime import datetime, date
from typing import List, Dict, Any


def generate_excel(
    data: List[Dict[str, Any]],
    output_path: str,
    fields: List[str],
) -> str:
    """Generate Excel file from extracted data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    # Header styling
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Build headers (source file + fields)
    headers = ["Source File"] + fields

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Data rows
    alt_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

    for row_idx, row_data in enumerate(data, 2):
        fill = alt_fill if row_idx % 2 == 0 else None

        # Source file
        cell = ws.cell(row=row_idx, column=1, value=row_data.get("_source_file", ""))
        cell.alignment = Alignment(vertical="center")
        if fill:
            cell.fill = fill

        for col_idx, field in enumerate(fields, 2):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(field, ""))
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if fill:
                cell.fill = fill

    # Auto-fit column widths
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_length = len(str(header))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, min(len(str(cell.value)), 50))
        ws.column_dimensions[col_letter].width = max_length + 4

    # Freeze header row
    ws.freeze_panes = "A2"

    # Row height for header
    ws.row_dimensions[1].height = 30

    wb.save(output_path)
    return output_path


def generate_excel_bytes(
    data: List[Dict[str, Any]],
    fields: List[str],
) -> bytes:
    """Generate Excel workbook and return raw bytes (for in-memory upload)."""
    buf = io.BytesIO()
    # Reuse generate_excel logic by saving to a BytesIO buffer
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = ["Source File"] + fields
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    alt_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    for row_idx, row_data in enumerate(data, 2):
        fill = alt_fill if row_idx % 2 == 0 else None
        cell = ws.cell(row=row_idx, column=1, value=row_data.get("_source_file", ""))
        cell.alignment = Alignment(vertical="center")
        if fill:
            cell.fill = fill
        for col_idx, field in enumerate(fields, 2):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(field, ""))
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if fill:
                cell.fill = fill

    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_length = len(str(header))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, min(len(str(cell.value)), 50))
        ws.column_dimensions[col_letter].width = max_length + 4

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    wb.save(buf)
    return buf.getvalue()


def append_to_excel_bytes(
    existing_bytes: bytes,
    new_rows: List[Dict[str, Any]],
    fields: List[str],
) -> bytes:
    """Load an existing Excel workbook from bytes, append new data rows, return bytes."""
    buf = io.BytesIO(existing_bytes)
    wb = openpyxl.load_workbook(buf)
    ws = wb.active

    # Read existing header row → column index map
    existing_headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    header_map = {str(h): idx + 1 for idx, h in enumerate(existing_headers) if h}

    alt_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

    for row_data in new_rows:
        next_row = ws.max_row + 1
        fill = alt_fill if next_row % 2 == 0 else None

        cell = ws.cell(row=next_row, column=1, value=row_data.get("_source_file", ""))
        cell.alignment = Alignment(vertical="center")
        if fill:
            cell.fill = fill

        for field in fields:
            col_idx = header_map.get(field)
            if col_idx is None:
                continue
            cell = ws.cell(row=next_row, column=col_idx, value=row_data.get(field, ""))
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if fill:
                cell.fill = fill

    out_buf = io.BytesIO()
    wb.save(out_buf)
    return out_buf.getvalue()


def append_to_csv_bytes(
    existing_bytes: bytes,
    new_rows: List[Dict[str, Any]],
    fields: List[str],
) -> bytes:
    """Append new data rows to existing CSV bytes."""
    existing_text = existing_bytes.decode("utf-8-sig").rstrip("\n")
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=["Source File"] + fields, extrasaction="ignore")
    for row_data in new_rows:
        row = {"Source File": row_data.get("_source_file", "")}
        for field in fields:
            row[field] = row_data.get(field, "")
        writer.writerow(row)
    return (existing_text + "\n" + buf.getvalue()).encode("utf-8")


def generate_csv_bytes(
    data: List[Dict[str, Any]],
    fields: List[str],
) -> bytes:
    """Generate CSV and return raw bytes (for in-memory upload)."""
    headers = ["Source File"] + fields
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    for row_data in data:
        row = {"Source File": row_data.get("_source_file", "")}
        for field in fields:
            row[field] = row_data.get(field, "")
        writer.writerow(row)
    return buf.getvalue().encode("utf-8")


def read_headers_from_bytes(data: bytes, fmt: str) -> List[str]:
    """Return the header row (first row) from an existing xlsx or csv spreadsheet."""
    if fmt == "csv":
        text = data.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        try:
            return next(reader)
        except StopIteration:
            return []
    else:
        buf = io.BytesIO(data)
        wb = openpyxl.load_workbook(buf, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            return [str(c) if c is not None else "" for c in row]
        return []


def generate_csv(
    data: List[Dict[str, Any]],
    output_path: str,
    fields: List[str],
) -> str:
    """Generate CSV file from extracted data."""
    headers = ["Source File"] + fields

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=headers,
            extrasaction="ignore",
        )
        writer.writeheader()

        for row_data in data:
            row = {"Source File": row_data.get("_source_file", "")}
            for field in fields:
                row[field] = row_data.get(field, "")
            writer.writerow(row)

    return output_path


# ── Accounting format helpers ──────────────────────────────────────────────────

def _clean_amount(raw: str) -> str:
    """Normalize an amount string: strip $, commas, handle parentheses for negatives."""
    if not raw:
        return "0.00"
    s = str(raw).strip()
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1]
    s = s.replace("$", "").replace(",", "").strip()
    if not s:
        return "0.00"
    try:
        val = float(s)
        if neg:
            val = -abs(val)
        return f"{val:.2f}"
    except ValueError:
        return "0.00"


_DATE_FORMATS = [
    "%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d", "%Y/%m/%d",
    "%m/%d/%y", "%m-%d-%y", "%d/%m/%Y", "%d-%m-%Y",
    "%B %d, %Y", "%b %d, %Y", "%d %B %Y", "%d %b %Y",
    "%B %d %Y", "%b %d %Y",
]


def _parse_date(raw: str) -> date | None:
    if not raw:
        return None
    s = str(raw).strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    digits = re.sub(r"[^\d]", "", s)
    if len(digits) == 8:
        try:
            return datetime.strptime(digits, "%Y%m%d").date()
        except ValueError:
            pass
        try:
            return datetime.strptime(digits, "%m%d%Y").date()
        except ValueError:
            pass
    return None


def generate_quickbooks_csv_bytes(data: List[Dict[str, Any]]) -> bytes:
    """Generate QuickBooks Online compatible CSV: Date, Description, Amount."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Date", "Description", "Amount"])
    for row in data:
        if row.get("_error"):
            continue
        parsed = _parse_date(row.get("Date", ""))
        if not parsed:
            continue
        date_str = parsed.strftime("%m/%d/%Y")
        desc = (row.get("Description", "") or "").replace(",", " ").strip()[:255]
        amount = _clean_amount(row.get("Amount", ""))
        if desc:
            writer.writerow([date_str, desc, amount])
    return buf.getvalue().encode("utf-8")


def generate_qbo_bytes(data: List[Dict[str, Any]]) -> bytes:
    """Generate QBO/OFX file (Web Connect format) for QuickBooks Desktop / Xero."""
    transactions = []
    dates = []
    for row in data:
        if row.get("_error"):
            continue
        parsed = _parse_date(row.get("Date", ""))
        if not parsed:
            continue
        dates.append(parsed)
        amount = _clean_amount(row.get("Amount", ""))
        desc = (row.get("Description", "") or "").strip()[:32]
        transactions.append((parsed, amount, desc))

    if not transactions:
        dt_start = dt_end = datetime.now().strftime("%Y%m%d")
    else:
        dt_start = min(dates).strftime("%Y%m%d")
        dt_end = max(dates).strftime("%Y%m%d")

    now_str = datetime.now().strftime("%Y%m%d%H%M%S")

    header = (
        "OFXHEADER:100\n"
        "DATA:OFXSGML\n"
        "VERSION:102\n"
        "SECURITY:NONE\n"
        "ENCODING:USASCII\n"
        "CHARSET:1252\n"
        "COMPRESSION:NONE\n"
        "OLDFILEUID:NONE\n"
        "NEWFILEUID:NONE\n"
        "\n"
        "<OFX>\n"
        "<SIGNONMSGSRSV1>\n"
        "<SONRS>\n"
        f"<STATUS><CODE>0<SEVERITY>INFO</STATUS>\n"
        f"<DTSERVER>{now_str}\n"
        "<LANGUAGE>ENG\n"
        "</SONRS>\n"
        "</SIGNONMSGSRSV1>\n"
        "<BANKMSGSRSV1>\n"
        "<STMTTRNRS>\n"
        "<TRNUID>0\n"
        f"<STATUS><CODE>0<SEVERITY>INFO</STATUS>\n"
        "<STMTRS>\n"
        "<CURDEF>USD\n"
        "<BANKACCTFROM>\n"
        "<BANKID>000000000\n"
        "<ACCTID>000000000\n"
        "<ACCTTYPE>CHECKING\n"
        "</BANKACCTFROM>\n"
        "<BANKTRANLIST>\n"
        f"<DTSTART>{dt_start}\n"
        f"<DTEND>{dt_end}\n"
    )

    txn_blocks = []
    for i, (txn_date, amount, desc) in enumerate(transactions):
        trntype = "CREDIT" if float(amount) >= 0 else "DEBIT"
        dtposted = txn_date.strftime("%Y%m%d")
        fitid = f"{dtposted}{i + 1:05d}"
        txn_blocks.append(
            f"<STMTTRN>\n"
            f"<TRNTYPE>{trntype}\n"
            f"<DTPOSTED>{dtposted}\n"
            f"<TRNAMT>{amount}\n"
            f"<FITID>{fitid}\n"
            f"<NAME>{desc}\n"
            f"</STMTTRN>\n"
        )

    total = sum(float(_clean_amount(r.get("Amount", ""))) for r in data if not r.get("_error"))
    bal_str = f"{total:.2f}"

    footer = (
        "</BANKTRANLIST>\n"
        "<LEDGERBAL>\n"
        f"<BALAMT>{bal_str}\n"
        f"<DTASOF>{dt_end}\n"
        "</LEDGERBAL>\n"
        "</STMTRS>\n"
        "</STMTTRNRS>\n"
        "</BANKMSGSRSV1>\n"
        "</OFX>\n"
    )

    return (header + "".join(txn_blocks) + footer).encode("ascii", errors="replace")
