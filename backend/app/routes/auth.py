import io
import json
import logging
import secrets as _secrets

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from pydantic import BaseModel
from app.config import settings
from app.database import get_db
from app.models.user import User
from app.services.auth_service import (
    create_access_token,
    get_or_create_user,
    verify_google_access_token,
    verify_microsoft_access_token,
)

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/auth", tags=["auth"])


class GoogleAuthRequest(BaseModel):
    access_token: str


class MicrosoftAuthRequest(BaseModel):
    access_token: str


class AuthResponse(BaseModel):
    access_token: str
    user: dict


class DevLoginRequest(BaseModel):
    secret: str


class DevSetUsageRequest(BaseModel):
    secret: str
    subscription_tier: str  # free | starter | pro | business
    pages_used_this_period: int
    overage_pages_this_period: int = 0


@router.post("/dev-set-usage", include_in_schema=False)
async def dev_set_usage(body: DevSetUsageRequest, db: AsyncSession = Depends(get_db)):
    """Dev-only: force the dev-login user's tier + usage for live test scenarios.
    Gated by DEV_LOGIN_SECRET so it mirrors /dev-login's trust boundary.
    """
    dev_secret = (settings.dev_login_secret or "").strip()
    if not dev_secret:
        raise HTTPException(status_code=404, detail="Not found")
    if not _secrets.compare_digest(body.secret, dev_secret):
        raise HTTPException(status_code=401, detail="Invalid secret")
    if body.subscription_tier not in ("free", "starter", "pro", "business"):
        raise HTTPException(status_code=400, detail="Invalid tier")
    if body.pages_used_this_period < 0 or body.overage_pages_this_period < 0:
        raise HTTPException(status_code=400, detail="Negative usage not allowed")

    user_id = (settings.dev_login_user_id or "").strip()
    if not user_id:
        raise HTTPException(status_code=503, detail="Dev user not configured")

    result = await db.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    user.subscription_tier = body.subscription_tier
    user.pages_used_this_period = body.pages_used_this_period
    user.overage_pages_this_period = body.overage_pages_this_period
    await db.commit()

    try:
        from app.cache import cache_del_user
        await cache_del_user(str(user.id))
    except Exception:
        pass

    return {
        "ok": True,
        "subscription_tier": user.subscription_tier,
        "pages_used_this_period": user.pages_used_this_period,
        "overage_pages_this_period": user.overage_pages_this_period,
    }


@router.post("/dev-llm-judge", include_in_schema=False)
async def dev_llm_judge(
    secret: str = Form(...),
    expectation: str = Form(...),
    file: UploadFile = File(...),
):
    """Dev-only: extract text from a PDF or xlsx and ask an LLM whether it
    satisfies ``expectation``. Used by Playwright live tests to replace brittle
    length-only output validation with semantic correctness checks.

    Returns ``{"verdict": "pass"|"fail", "reasoning": "..."}``. The client is
    responsible for treating ``verdict != "pass"`` as a test failure.
    """
    dev_secret = (settings.dev_login_secret or "").strip()
    if not dev_secret:
        raise HTTPException(status_code=404, detail="Not found")
    if not _secrets.compare_digest(secret, dev_secret):
        raise HTTPException(status_code=401, detail="Invalid secret")

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Empty file")

    filename = (file.filename or "").lower()
    content_type = (file.content_type or "").lower()

    # ── Extract text from the artifact ─────────────────────────────────────
    text: str = ""
    try:
        if filename.endswith(".pdf") or "pdf" in content_type:
            from pypdf import PdfReader
            reader = PdfReader(io.BytesIO(raw))
            pages_text = []
            for page in reader.pages[:20]:  # cap at 20 pages to keep the prompt small
                try:
                    pages_text.append(page.extract_text() or "")
                except Exception as exc:
                    logger.warning("pypdf extract failed on a page: %s", exc)
            text = "\n\n--- PAGE BREAK ---\n\n".join(pages_text)
        elif filename.endswith(".xlsx") or "sheet" in content_type or "excel" in content_type:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
            lines = []
            for ws in wb.worksheets[:5]:  # cap at 5 sheets
                lines.append(f"[Sheet: {ws.title}]")
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= 200:
                        lines.append(f"... ({i} rows truncated)")
                        break
                    lines.append("\t".join("" if v is None else str(v) for v in row))
            text = "\n".join(lines)
        else:
            raise HTTPException(status_code=400, detail=f"Unsupported file type: {filename} / {content_type}")
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("llm-judge text extraction failed")
        raise HTTPException(status_code=400, detail=f"Could not parse file: {exc}")

    if not text.strip():
        return {
            "verdict": "fail",
            "reasoning": "No extractable text found in the artifact (possibly scanned/image-only or corrupted).",
            "extracted_chars": 0,
        }

    # ── Ask an LLM to judge ────────────────────────────────────────────────
    snippet = text[:12000]  # keep prompt bounded
    system = (
        "You are a strict QA judge validating test artifacts produced by a document-processing "
        "pipeline. You will be given the extracted textual content of a PDF or spreadsheet and a "
        "natural-language expectation. Return STRICT JSON with two keys: "
        '{"verdict": "pass" | "fail", "reasoning": "<one or two sentences>"}. '
        "Return 'pass' only if the content clearly satisfies the expectation. "
        "Return 'fail' if the content is empty, unrelated, garbled, or only partially present. "
        "Do NOT wrap the JSON in markdown. Output JSON only."
    )
    user_prompt = (
        f"EXPECTATION:\n{expectation}\n\n"
        f"EXTRACTED CONTENT (first 12k chars):\n{snippet}\n\n"
        "Return JSON now."
    )
    try:
        from app.services.llm_router import _get_openai_client
        client = _get_openai_client()
        resp = await client.chat.completions.create(
            model=settings.llm_openai_fallback_model or "gpt-4.1-mini",
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": user_prompt},
            ],
            temperature=0.0,
            max_tokens=250,
            response_format={"type": "json_object"},
        )
        raw_out = (resp.choices[0].message.content or "").strip()
        parsed = json.loads(raw_out)
        verdict = str(parsed.get("verdict", "fail")).lower()
        if verdict not in ("pass", "fail"):
            verdict = "fail"
        return {
            "verdict": verdict,
            "reasoning": str(parsed.get("reasoning", "") or "")[:500],
            "extracted_chars": len(text),
        }
    except Exception as exc:
        logger.exception("llm-judge OpenAI call failed")
        raise HTTPException(status_code=502, detail=f"LLM judge failed: {exc}")


@router.post("/dev-login")
async def dev_login(body: DevLoginRequest, db: AsyncSession = Depends(get_db)):
    """Bypass OAuth for dev/test. Disabled unless DEV_LOGIN_SECRET is set in env."""
    dev_secret = (settings.dev_login_secret or "").strip()
    if not dev_secret:
        raise HTTPException(status_code=404, detail="Not found")
    if not _secrets.compare_digest(body.secret, dev_secret):
        raise HTTPException(status_code=401, detail="Invalid secret")

    user_id = (settings.dev_login_user_id or "").strip()
    if not user_id:
        raise HTTPException(status_code=503, detail="Dev login user not configured")

    result = await db.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    if not user or not user.is_active:
        raise HTTPException(status_code=404, detail="User not found")

    token = create_access_token(user.id)
    period_end = user.current_period_end.isoformat() if user.current_period_end else None
    return {
        "access_token": token,
        "user": {
            "id": user.id,
            "email": user.email,
            "name": user.name,
            "picture": user.picture,
            "balance": user.balance,
            "has_card": bool(user.stripe_payment_method_id),
            "subscription_tier": user.subscription_tier or "free",
            "subscription_status": user.subscription_status or "active",
            "pages_used_this_period": user.pages_used_this_period or 0,
            "current_period_end": period_end,
        },
    }


@router.post("/google", response_model=AuthResponse)
async def google_auth(request: Request, body: GoogleAuthRequest, db: AsyncSession = Depends(get_db)):
    """Authenticate with Google OAuth access token."""
    client_ip = request.client.host if request.client else "-"
    token_preview = body.access_token[:12] + "…" if len(body.access_token) > 12 else body.access_token

    logger.info("Google login attempt from %s (token: %s)", client_ip, token_preview)

    try:
        google_user = await verify_google_access_token(body.access_token)
    except Exception as e:
        logger.warning("Google token verification failed from %s: %s", client_ip, str(e))
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail=f"Invalid Google token: {str(e)}",
        )

    email = google_user.get("email", "<no-email>")
    name = google_user.get("name", "<no-name>")
    logger.info("Google token verified — email=%s name=%s ip=%s", email, name, client_ip)

    user = await get_or_create_user(db, google_user, provider="google")
    token = create_access_token(user.id)

    logger.info("Login successful — user_id=%s email=%s balance=$%.6f ip=%s", user.id, user.email, user.balance, client_ip)

    period_end = user.current_period_end.isoformat() if user.current_period_end else None
    return {
        "access_token": token,
        "user": {
            "id": user.id,
            "email": user.email,
            "name": user.name,
            "picture": user.picture,
            "balance": user.balance,
            "has_card": bool(user.stripe_payment_method_id),
            "subscription_tier": user.subscription_tier or "free",
            "subscription_status": user.subscription_status or "active",
            "pages_used_this_period": user.pages_used_this_period or 0,
            "current_period_end": period_end,
        },
    }


@router.post("/microsoft", response_model=AuthResponse)
async def microsoft_auth(request: Request, body: MicrosoftAuthRequest, db: AsyncSession = Depends(get_db)):
    """Authenticate with Microsoft OAuth access token."""
    client_ip = request.client.host if request.client else "-"
    token_preview = body.access_token[:12] + "…" if len(body.access_token) > 12 else body.access_token

    logger.info("Microsoft login attempt from %s (token: %s)", client_ip, token_preview)

    try:
        ms_user = await verify_microsoft_access_token(body.access_token)
    except Exception as e:
        logger.warning("Microsoft token verification failed from %s: %s", client_ip, str(e))
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail=f"Invalid Microsoft token: {str(e)}",
        )

    email = ms_user.get("email", "<no-email>")
    name = ms_user.get("name", "<no-name>")
    logger.info("Microsoft token verified — email=%s name=%s ip=%s", email, name, client_ip)

    user = await get_or_create_user(db, ms_user, provider="microsoft")
    token = create_access_token(user.id)

    logger.info("Login successful — user_id=%s email=%s balance=$%.6f ip=%s", user.id, user.email, user.balance, client_ip)

    period_end = user.current_period_end.isoformat() if user.current_period_end else None
    return {
        "access_token": token,
        "user": {
            "id": user.id,
            "email": user.email,
            "name": user.name,
            "picture": user.picture,
            "balance": user.balance,
            "has_card": bool(user.stripe_payment_method_id),
            "subscription_tier": user.subscription_tier or "free",
            "subscription_status": user.subscription_status or "active",
            "pages_used_this_period": user.pages_used_this_period or 0,
            "current_period_end": period_end,
        },
    }
