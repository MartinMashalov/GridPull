"""
Production e2e test: email ingest → inbox group → SOV extraction → baseline update.

Steps:
  1. Send emails with SOV PDF attachments to documents@gridpull.com
  2. Poll the production API until the docs appear in the inbox
  3. Trigger SOV extraction on the group
  4. Poll until the job is done
  5. Download the result Excel and report fill rates

Usage:
    cd backend
    python tests/_test_email_ingest_e2e.py
"""
from __future__ import annotations

import asyncio
import base64
import io
import json
import smtplib
import sys
import time
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import httpx
import openpyxl
from jose import jwt

_BACKEND = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_BACKEND))

import os
try:
    from dotenv import load_dotenv; load_dotenv(_BACKEND / ".env")
except ImportError:
    pass
os.environ.pop("STRIPE_PRODUCT_ID", None)

from app.config import settings

# ── Config ────────────────────────────────────────────────────────────────────

GMAIL_USER     = os.getenv("GMAIL_IMAP_EMAIL", "")
GMAIL_PASSWORD = os.getenv("GMAIL_IMAP_PASSWORD", "")   # App Password
INGEST_EMAIL   = settings.ingest_universal_email
API_BASE       = "https://gridpull.com/api"
JWT_SECRET     = settings.jwt_secret_key
JWT_ALGORITHM  = settings.jwt_algorithm

# PDFs to send — pick the 3 richest SOV docs for maximum field coverage
TEST_FILES = [
    _BACKEND / "test_files" / "01_property_appraisal_report_25_buildings.pdf",
    _BACKEND / "test_files" / "02_carrier_property_schedule_50_locations.pdf",
    _BACKEND / "test_files" / "05_prior_year_sov_30_locations.pdf",
]

# Unique sender tag so we can identify our test batch in the inbox
import secrets as _sec
TEST_TAG      = _sec.token_hex(4)
TEST_SENDER   = f"sov-test-{TEST_TAG}@gridpull-test.com"
TEST_SUBJECT  = f"SOV Test Batch {TEST_TAG}"

GREEN  = "\033[92m"; YELLOW = "\033[93m"; RED = "\033[91m"
BOLD   = "\033[1m";  RESET  = "\033[0m";  CYAN = "\033[96m"

# ── Field set (exact frontend defaults) ──────────────────────────────────────

from tests.test_sov_e2e import FIELD_SETS
FIELDS      = FIELD_SETS["frontend_defaults"]
FIELD_NAMES = [f["name"] for f in FIELDS]

# ── Helpers ───────────────────────────────────────────────────────────────────

def _make_jwt(user_id: str) -> str:
    now = int(time.time())
    return jwt.encode(
        {"sub": user_id, "iat": now, "exp": now + 3600},
        JWT_SECRET, algorithm=JWT_ALGORITHM,
    )


async def _get_user_id() -> str:
    """Query production DB for martin.mashalov@gmail.com user ID."""
    import asyncpg
    db_url = str(settings.database_url).replace("postgresql+asyncpg://", "postgresql://")
    conn = await asyncpg.connect(db_url)
    try:
        row = await conn.fetchrow("SELECT id FROM users WHERE email = $1", GMAIL_USER)
        if not row:
            raise RuntimeError(f"User {GMAIL_USER} not found in production DB")
        return str(row["id"])
    finally:
        await conn.close()


def _send_emails() -> int:
    """Send one email per PDF to the ingest address. Returns number of emails sent."""
    sent = 0
    smtp = smtplib.SMTP_SSL("smtp.gmail.com", 465)
    smtp.login(GMAIL_USER, GMAIL_PASSWORD)

    for pdf_path in TEST_FILES:
        if not pdf_path.exists():
            print(f"  {YELLOW}Skipping missing file: {pdf_path.name}{RESET}")
            continue

        msg = MIMEMultipart()
        msg["From"]    = GMAIL_USER
        msg["To"]      = INGEST_EMAIL
        msg["Subject"] = f"{TEST_SUBJECT} — {pdf_path.name}"
        # Custom header so webhook can group by our test tag
        msg["X-Test-Tag"] = TEST_TAG
        msg.attach(MIMEText(f"SOV test batch {TEST_TAG}\nFile: {pdf_path.name}", "plain"))

        with open(pdf_path, "rb") as fh:
            part = MIMEApplication(fh.read(), Name=pdf_path.name)
        part["Content-Disposition"] = f'attachment; filename="{pdf_path.name}"'
        msg.attach(part)

        smtp.sendmail(GMAIL_USER, INGEST_EMAIL, msg.as_string())
        print(f"  {GREEN}Sent:{RESET} {pdf_path.name} → {INGEST_EMAIL}")
        sent += 1
        time.sleep(0.5)

    smtp.quit()
    return sent


async def _upload_to_inbox(token: str) -> list[str]:
    """POST /ingest/inbox/upload — upload all test PDFs directly into the inbox."""
    headers = {"Authorization": f"Bearer {token}"}
    doc_ids: list[str] = []

    async with httpx.AsyncClient(timeout=60) as client:
        for pdf_path in TEST_FILES:
            if not pdf_path.exists():
                print(f"  {YELLOW}Skipping missing file: {pdf_path.name}{RESET}")
                continue
            data = pdf_path.read_bytes()
            files = [("files", (pdf_path.name, data, "application/pdf"))]
            form = {"sender_email": f"test-batch-{TEST_TAG}@gridpull-test.com", "sender_domain": "gridpull-test.com"}
            resp = await client.post(f"{API_BASE}/ingest/inbox/upload", headers=headers, files=files, data=form)
            if resp.status_code != 200:
                print(f"  {RED}Upload failed for {pdf_path.name}: {resp.status_code} {resp.text[:200]}{RESET}")
                continue
            result = resp.json()
            for u in result.get("uploaded", []):
                doc_ids.append(u["id"])
            print(f"  {GREEN}Uploaded:{RESET} {pdf_path.name} (id={result.get('uploaded', [{}])[0].get('id', '?')[:8]}...)")

    return doc_ids


async def _poll_inbox(token: str, expected_count: int, timeout: int = 120) -> list[str]:
    """Poll GET /ingest/inbox until we see our test documents. Returns doc IDs."""
    headers = {"Authorization": f"Bearer {token}"}
    deadline = time.time() + timeout

    async with httpx.AsyncClient(timeout=30) as client:
        while time.time() < deadline:
            resp = await client.get(f"{API_BASE}/ingest/inbox", headers=headers)
            if resp.status_code != 200:
                print(f"  {YELLOW}Inbox poll returned {resp.status_code}, retrying...{RESET}")
                await asyncio.sleep(5)
                continue

            data = resp.json()
            groups = data.get("groups", [])

            # Find our test group by sender matching our test tag
            our_docs: list[str] = []
            for group in groups:
                if TEST_TAG in (group.get("key") or ""):
                    our_docs = [d["id"] for d in group.get("documents", [])]
                    break

            print(f"  Inbox: {data.get('total_documents', 0)} total docs, our test group: {len(our_docs)}/{expected_count}")
            if len(our_docs) >= expected_count:
                return our_docs

            await asyncio.sleep(8)

    print(f"  {RED}Timeout waiting for inbox documents{RESET}")
    return []


async def _trigger_extraction(token: str, doc_ids: list[str]) -> str | None:
    """POST /ingest/inbox/extract. Returns job_id."""
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    payload = {
        "document_ids": doc_ids,
        "fields": FIELDS,
        "instructions": "",
        "format": "xlsx",
        "pipeline": "sov",
    }
    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.post(f"{API_BASE}/ingest/inbox/extract", headers=headers, json=payload)
    if resp.status_code != 200:
        print(f"  {RED}Extract request failed {resp.status_code}: {resp.text[:300]}{RESET}")
        return None
    return resp.json().get("job_id")


async def _poll_job(token: str, job_id: str, timeout: int = 600) -> dict | None:
    """Poll GET /documents/jobs/{job_id} until done."""
    headers = {"Authorization": f"Bearer {token}"}
    deadline = time.time() + timeout

    async with httpx.AsyncClient(timeout=30) as client:
        while time.time() < deadline:
            resp = await client.get(f"{API_BASE}/documents/job/{job_id}", headers=headers)
            if resp.status_code != 200:
                await asyncio.sleep(5)
                continue
            job = resp.json()
            status = job.get("status", "")
            pct = job.get("progress", 0)
            print(f"  Job {job_id[:8]}... status={status} progress={pct}%")
            if status in ("done", "complete", "completed", "failed", "error"):
                return job
            await asyncio.sleep(10)

    print(f"  {RED}Timeout waiting for job{RESET}")
    return None


async def _download_result(token: str, job_id: str) -> bytes | None:
    """GET /documents/download/{job_id}"""
    headers = {"Authorization": f"Bearer {token}"}
    async with httpx.AsyncClient(timeout=60) as client:
        resp = await client.get(f"{API_BASE}/documents/download/{job_id}", headers=headers)
    if resp.status_code != 200:
        print(f"  {RED}Download failed {resp.status_code}{RESET}")
        return None
    return resp.content


def _analyse_excel(data: bytes) -> None:
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    data_rows = ws.max_row - 1
    print(f"\n  {CYAN}Result spreadsheet:{RESET}")
    print(f"    Rows: {data_rows}")
    print(f"    Columns: {len(headers)}")

    # Per-field fill rates for our 46 target fields
    print(f"\n  {CYAN}Per-field fill rates:{RESET}")
    _EMPTY = {"", "null", "none", "n/a", "na", "-"}
    for fn in FIELD_NAMES:
        try:
            col = headers.index(fn) + 1
        except ValueError:
            print(f"    {RED}MISSING COLUMN: {fn}{RESET}")
            continue
        filled = sum(
            1 for r in range(2, ws.max_row + 1)
            if str(ws.cell(r, col).value or "").strip().lower() not in _EMPTY
        )
        pct = filled / data_rows if data_rows else 0
        color = GREEN if pct >= 0.7 else (YELLOW if pct >= 0.3 else RED)
        bar = "█" * int(pct * 10) + "░" * (10 - int(pct * 10))
        print(f"    {color}{bar} {pct:5.0%} {fn}{RESET}")

    # Overall fill rate
    total_cells = data_rows * len(FIELD_NAMES)
    filled_cells = sum(
        1 for fn in FIELD_NAMES
        for r in range(2, ws.max_row + 1)
        if (col := (headers.index(fn) + 1 if fn in headers else None))
        and str(ws.cell(r, col).value or "").strip().lower() not in _EMPTY
    )
    print(f"\n  Overall fill rate: {BOLD}{filled_cells/total_cells:.1%}{RESET} ({filled_cells}/{total_cells} cells)")


# ── Main ──────────────────────────────────────────────────────────────────────

async def main() -> None:
    print(f"\n{BOLD}{'█' * 72}{RESET}")
    print(f"{BOLD}  Production E2E: Email Ingest → SOV Extraction{RESET}")
    print(f"  Sender:    {GMAIL_USER}")
    print(f"  To:        {INGEST_EMAIL}")
    print(f"  Test tag:  {TEST_TAG}")
    print(f"  Files:     {[p.name for p in TEST_FILES if p.exists()]}")
    print(f"{BOLD}{'█' * 72}{RESET}\n")

    # Step 1: get user ID + JWT
    print(f"{CYAN}Step 1: Auth{RESET}")
    user_id = await _get_user_id()
    token = _make_jwt(user_id)
    print(f"  User ID: {user_id}")
    print(f"  JWT:     {token[:40]}...")

    # Step 2: upload files directly to the production inbox via API
    print(f"\n{CYAN}Step 2: Upload files to production inbox{RESET}")
    doc_ids = await _upload_to_inbox(token)
    if not doc_ids:
        print(f"{RED}Upload failed — aborting{RESET}")
        return
    print(f"  {GREEN}Uploaded {len(doc_ids)} documents to inbox{RESET}")

    # Step 4: trigger extraction
    print(f"\n{CYAN}Step 4: Trigger SOV extraction{RESET}")
    job_id = await _trigger_extraction(token, doc_ids)
    if not job_id:
        return
    print(f"  Job ID: {job_id}")

    # Step 5: poll job
    print(f"\n{CYAN}Step 5: Waiting for job to complete...{RESET}")
    job = await _poll_job(token, job_id, timeout=600)
    if not job:
        return
    print(f"  Status: {job.get('status')}")
    if job.get("status") in ("failed", "error"):
        print(f"  {RED}Job failed: {job.get('error_message', '')}{RESET}")
        return

    # Step 6: download + analyse
    print(f"\n{CYAN}Step 6: Download + analyse result{RESET}")
    result_bytes = await _download_result(token, job_id)
    if not result_bytes:
        return

    out_path = _BACKEND / "test_files" / f"email_ingest_result_{TEST_TAG}.xlsx"
    out_path.write_bytes(result_bytes)
    print(f"  Saved → {out_path.name}")
    _analyse_excel(result_bytes)

    print(f"\n{BOLD}{'█' * 72}{RESET}")
    print(f"{BOLD}  DONE{RESET}")
    print(f"{'█' * 72}{RESET}\n")


if __name__ == "__main__":
    asyncio.run(main())
