"""
Production-readiness stress test for the SOV extraction pipeline.

Runs 4 batteries of tests:

  A) Single-doc: original files 01-05 (25 / 50 / 15 / 35 / 30 rows)
     — verifies the single-doc OCR path still works after all recent changes.

  B) Consistency: run C07 (06+09, the previously garbled combo) 3× to confirm
     the address-rule fix holds across multiple LLM calls.

  C) Excel output inspection: generate actual .xlsx from C15 (all 4 docs, 20
     locations) and inspect every cell — no nulls in core columns, no dups,
     no garble, values look sane (TIV > 0, year 1900-2030, state 2-char).

  D) Address quality: inspect all 20 Street Address / City / State / Zip
     values from C01, C07, C11, C15 looking for truncation or garble.

Usage:
    cd backend
    python tests/test_sov_production_ready.py
"""
from __future__ import annotations

import asyncio
import io
import re
import sys
import time
from pathlib import Path
from typing import Any, Dict, List

import openpyxl

_BACKEND = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_BACKEND))

import os
try:
    from dotenv import load_dotenv
    load_dotenv(_BACKEND / ".env")
except ImportError:
    pass
os.environ.pop("STRIPE_PRODUCT_ID", None)

from app.config import settings
from app.services.extraction import extract_from_document, LLMUsage
from app.services.pdf_service import parse_pdf, combine_parsed_documents
from app.services.ocr_service import run_mistral_ocr
from app.services.sov.pipeline import _build_sections_from_ocr_pages
from app.services.spreadsheet_service import generate_excel_bytes

GREEN  = "\033[92m"
YELLOW = "\033[93m"
RED    = "\033[91m"
BOLD   = "\033[1m"
DIM    = "\033[2m"
RESET  = "\033[0m"
CYAN   = "\033[96m"

_EMPTY = {"", "null", "none", "n/a", "na", "-", "—", "unknown", "not found", "not available"}
_BAD_OCR = [
    "dban", "maysdby", "harb or", "st ree", "dbay", "dbos",
    # patterns from Mistral OCR garbling "San" city names into the street field
    "dsran", "stsan", "stden", "stchi", "sthou", "stpho", "stdal", "staus",
    # city name concatenated directly after street suffix with no space
    r"st[A-Z][a-z]{2,}",  # e.g. "StSan", "StDen"
]


def is_filled(v: Any) -> bool:
    return v is not None and str(v).strip().lower() not in _EMPTY


def fill_rate(rows: List[Dict], field_names: List[str]) -> float:
    total = len(rows) * len(field_names)
    return sum(1 for r in rows for fn in field_names if is_filled(r.get(fn))) / total if total else 0.0


def find_garble(rows: List[Dict], field_names: List[str]) -> List[str]:
    hits = []
    for r in rows:
        for fn in field_names:
            val = str(r.get(fn) or "")
            val_lower = val.lower()
            for pat in _BAD_OCR:
                if pat.startswith("r"):
                    # regex pattern
                    if re.search(pat[1:], val):
                        hits.append(f"Loc#{r.get('Loc #','?')} {fn}={val!r}")
                        break
                elif pat in val_lower:
                    hits.append(f"Loc#{r.get('Loc #','?')} {fn}={val!r}")
                    break
    return hits


def find_dups(rows: List[Dict]) -> List[str]:
    seen: Dict[str, int] = {}
    for r in rows:
        loc = str(r.get("Loc #", "") or "").strip()
        if loc:
            seen[loc] = seen.get(loc, 0) + 1
    return [k for k, v in seen.items() if v > 1]


# ──────────────────────────────────────────────────────────────────────────────
# Field sets
# ──────────────────────────────────────────────────────────────────────────────

FRONTEND_DEFAULTS = [
    {"name": "Loc #",                        "description": "Extract the location identifier exactly as shown for each schedule row (for example: 1, 01, A1). Keep letters, symbols, and leading zeros exactly as printed."},
    {"name": "Bldg #",                       "description": "Extract the building number for the location exactly as shown in the schedule. Do not infer or renumber buildings; copy the literal value from the row."},
    {"name": "Location Name",                "description": "Extract the site or building name used by underwriting (for example: MB1, North Warehouse). If no distinct name is present, derive from the location identifier."},
    {"name": "Occupancy/Exposure",           "description": "Extract the occupancy/exposure classification text exactly as presented."},
    {"name": "Street Address",               "description": "Extract the street address line for the insured premises. Keep suite/unit/building details when present, but do not include city/state/zip in this field unless the source combines them into one cell."},
    {"name": "City",                         "description": "Extract the city for the insured location exactly as listed in the schedule row."},
    {"name": "State",                        "description": "Extract the state value exactly as shown (postal abbreviation preferred)."},
    {"name": "Zip",                          "description": "Extract the ZIP/postal code exactly as shown, including ZIP+4 when present."},
    {"name": "County",                       "description": "Extract the county value exactly as shown."},
    {"name": "Construction Type",            "description": "Extract the construction class/type used in underwriting (for example: Frame, Joisted Masonry, Non-Combustible). Keep the schedule wording as-is."},
    {"name": "ISO Construction Code",        "description": "Extract the ISO construction code exactly as shown (for example: F, JM, NC, 1-6)."},
    {"name": "Building Values",              "description": "Extract the building limit/value amount for the row. Keep the currency format shown in the source unless the source is plain numeric."},
    {"name": "Contents/BPP Values",          "description": "Extract the contents/business personal property value for the row."},
    {"name": "Business Income Values",       "description": "Extract the business income/time element value exactly as shown for the location row."},
    {"name": "Machinery & Equipment Values", "description": "Extract the machinery and equipment value exactly as shown for the row."},
    {"name": "Other Property Values",        "description": "Extract the other property value exactly as presented for the row."},
    {"name": "Total Insurable Value (TIV)",  "description": "Extract the total insurable value for this row. In property schedules this column is typically labeled 'TIV' — map that column directly to this field."},
    {"name": "Square Ft.",                   "description": "Extract the insured area in square feet exactly as shown."},
    {"name": "Cost Per Square Ft.",          "description": "Extract cost per square foot exactly as shown."},
    {"name": "Year Built",                   "description": "Extract original year built for the building row."},
    {"name": "Roof Update",                  "description": "Extract the roof update year or indicator exactly as shown."},
    {"name": "Wiring Update",                "description": "Extract the wiring update year or indicator exactly as shown."},
    {"name": "HVAC Update",                  "description": "Extract the HVAC update year or indicator exactly as shown."},
    {"name": "Plumbing Update",              "description": "Extract the plumbing update year or indicator exactly as shown."},
    {"name": "% Occupied",                   "description": "Extract occupancy percentage exactly as shown."},
    {"name": "Sprinklered",                  "description": "Extract sprinkler status exactly as shown."},
    {"name": "% Sprinklered",                "description": "Extract sprinkler percentage exactly as shown."},
    {"name": "ISO Protection Class",         "description": "Extract ISO protection class exactly as shown in the row."},
    {"name": "Fire Alarm",                   "description": "Extract fire alarm indicator exactly as shown."},
    {"name": "Burglar Alarm",                "description": "Extract burglar alarm indicator exactly as shown."},
    {"name": "Smoke Detectors",              "description": "Extract smoke detector indicator/status exactly as shown."},
    {"name": "# of Stories",                "description": "Extract number of stories exactly as shown."},
    {"name": "# of Units",                  "description": "Extract number of units exactly as shown in the row."},
    {"name": "Type of Wiring",              "description": "Extract type of wiring code or text exactly as shown."},
    {"name": "% Subsidized",               "description": "Extract subsidized occupancy percentage exactly as shown."},
    {"name": "% Student Housing",          "description": "Extract student housing percentage exactly as shown."},
    {"name": "% Elderly Housing",          "description": "Extract elderly housing percentage exactly as shown."},
    {"name": "Roof Type/Frame",            "description": "Extract roof type/frame value exactly as shown."},
    {"name": "Roof Shape",                 "description": "Extract roof shape code/text exactly as shown."},
    {"name": "Flood Zone",                 "description": "Extract FEMA flood zone exactly as shown."},
    {"name": "EQ Zone",                    "description": "Extract earthquake zone code/classification exactly as shown."},
    {"name": "Distance to Salt Water/Coast","description": "Extract the distance-to-coast value exactly as shown."},
    {"name": "Property Owned or Managed",  "description": "Extract owned/managed indicator exactly as shown."},
    {"name": "Bldg Maintenance",           "description": "Extract building maintenance indicator/class exactly as shown."},
    {"name": "Basement",                   "description": "Extract basement indicator exactly as shown."},
    {"name": "Predominant Exterior Wall / Cladding","description": "Extract predominant exterior wall/cladding material exactly as shown."},
]

ABBREVIATED = [
    {"name": "Loc #",         "description": "Location or property number/ID"},
    {"name": "Street Address","description": "Street address of the property"},
    {"name": "City",          "description": "City"},
    {"name": "State",         "description": "State (2-letter code)"},
    {"name": "Zip",           "description": "Zip code"},
    {"name": "Construction Type","description": "Construction type (e.g. frame, masonry)"},
    {"name": "Year Built",    "description": "Year the building was constructed"},
    {"name": "Building Values","description": "Building insured value in dollars"},
    {"name": "Total Insurable Value (TIV)","description": "Total insured value in dollars"},
]

VEHICLE_FIELDS = [
    {"name": "Vehicle #",     "description": "Vehicle or unit number"},
    {"name": "Year",          "description": "Vehicle model year"},
    {"name": "Make",          "description": "Vehicle manufacturer"},
    {"name": "Model",         "description": "Vehicle model name"},
    {"name": "VIN",           "description": "Vehicle identification number"},
    {"name": "Body Type",     "description": "Vehicle body type (e.g. Sedan, Pickup, Van)"},
    {"name": "Cost New",      "description": "Original cost or stated value of the vehicle"},
    {"name": "Garaging City", "description": "City where the vehicle is garaged"},
    {"name": "Garaging State","description": "State where the vehicle is garaged"},
    {"name": "Garaging Zip",  "description": "Zip code where the vehicle is garaged"},
]

FILES_05 = {
    "01": (_BACKEND / "test_files" / "01_property_appraisal_report_25_buildings.pdf", 25, FRONTEND_DEFAULTS),
    "02": (_BACKEND / "test_files" / "02_carrier_property_schedule_50_locations.pdf", 50, ABBREVIATED),
    "03": (_BACKEND / "test_files" / "03_package_policy_dec_pages_15_locations.pdf", 15, FRONTEND_DEFAULTS),
    "04": (_BACKEND / "test_files" / "04_vehicle_schedule_35_vehicles.pdf",           35, VEHICLE_FIELDS),
    "05": (_BACKEND / "test_files" / "05_prior_year_sov_30_locations.pdf",            30, ABBREVIATED),
}

FILES_0609 = {
    "06": _BACKEND / "test_files" / "06_full_sov_20_locations.pdf",
    "07": _BACKEND / "test_files" / "07_customer_intake_form.pdf",
    "08": _BACKEND / "test_files" / "08_appraisal_supplement.pdf",
    "09": _BACKEND / "test_files" / "09_email_thread_updates.pdf",
}

_NO_OCR  = {".eml", ".emlx", ".msg", ".html", ".htm"}
_SHEET   = {".xlsx", ".xls", ".xlsm", ".csv"}


async def _parse_ocr(path: Path) -> Any:
    """Mirror job_processor._parse_only: OCR only for scanned docs or dense_tables/mixed hints."""
    p = await asyncio.to_thread(parse_pdf, str(path), path.name)
    needs_ocr = p.is_scanned or p.doc_type_hint in ("dense_tables", "mixed")
    if (settings.mistral_api_key
            and needs_ocr
            and not any(path.suffix.lower() == e for e in _NO_OCR | _SHEET)):
        try:
            ocr = await run_mistral_ocr(str(path), settings.mistral_api_key, max_pages=50)
            secs = _build_sections_from_ocr_pages(ocr.pages)
            txt = "\n\n".join(s.content for s in secs).strip()
            if txt:
                p.content_text = txt
                print(f"    [OCR] {path.name} hint={p.doc_type_hint}: {ocr.page_count}p → {len(txt):,}c")
        except Exception as exc:
            print(f"    {YELLOW}OCR failed for {path.name}: {exc}{RESET}")
    else:
        print(f"    [liteparse] {path.name} hint={p.doc_type_hint} scanned={p.is_scanned}")
    return p


def _ok(cond: bool, msg: str) -> str:
    return f"{GREEN}✓{RESET} {msg}" if cond else f"{RED}✗ {msg}{RESET}"


# ══════════════════════════════════════════════════════════════════════════════
# BATTERY A — original files 01-05 (single-doc)
# ══════════════════════════════════════════════════════════════════════════════

async def battery_a() -> List[Dict]:
    print(f"\n{BOLD}{'█'*72}{RESET}")
    print(f"{BOLD}  BATTERY A — Single-doc: files 01-05{RESET}")
    print(f"{BOLD}{'█'*72}{RESET}")
    results = []
    for key, (path, expected, fields) in FILES_05.items():
        field_names = [f["name"] for f in fields]
        print(f"\n{BOLD}  A{key}: {path.name} (expect {expected} rows){RESET}")
        if not path.exists():
            print(f"  {RED}MISSING{RESET}"); continue
        parsed = await asyncio.to_thread(parse_pdf, str(path), path.name)
        print(f"  Parsed: pages={parsed.page_count} tables={len(parsed.tables)} hint={parsed.doc_type_hint} scanned={parsed.is_scanned}")
        usage = LLMUsage()
        t0 = time.perf_counter()
        rows = await extract_from_document(parsed, fields, usage, force_sov=True)
        elapsed = time.perf_counter() - t0
        real = [r for r in rows if not r.get("_error")]
        fr = fill_rate(real, field_names)
        garble = find_garble(real, field_names)
        dups = find_dups(real)
        count_ok = abs(len(real) - expected) <= max(3, int(expected * 0.15))
        fill_ok  = fr >= 0.40
        passed   = count_ok and fill_ok and not garble and not dups
        print(f"  {_ok(count_ok, f'rows={len(real)}/{expected}')}")
        print(f"  {_ok(fill_ok, f'fill={fr:.1%}')}")
        print(f"  {_ok(not dups, 'no duplicate IDs')}")
        print(f"  {_ok(not garble, 'no garbled text')}")
        print(f"  t={elapsed:.0f}s  cost=${usage.cost_usd:.4f}")
        # Sample rows
        print(f"  {DIM}First 3:{RESET}")
        id_field = field_names[0]
        addr_fields = [fn for fn in ["Street Address", "City", "State", "Garaging City", "Garaging State"] if fn in field_names]
        val_field = next((fn for fn in ["Total Insurable Value (TIV)", "Cost New", "TIV", "Building Values"] if fn in field_names), None)
        for r in real[:3]:
            loc = r.get(id_field)
            addr = " | ".join(str(r.get(fn) or "-") for fn in addr_fields[:3])
            val = r.get(val_field) if val_field else "-"
            print(f"    {id_field}={loc!r}  {addr}  {val_field}={val!r}")
        verdict = f"{GREEN}PASS{RESET}" if passed else f"{RED}FAIL{RESET}"
        print(f"  Verdict: {verdict}")
        results.append({"key": f"A{key}", "pass": passed, "rows": len(real), "expected": expected, "fill": fr, "elapsed": elapsed, "cost": usage.cost_usd})
    return results


# ══════════════════════════════════════════════════════════════════════════════
# BATTERY B — consistency: run C07 (06+09) three times
# ══════════════════════════════════════════════════════════════════════════════

async def battery_b() -> List[Dict]:
    print(f"\n{BOLD}{'█'*72}{RESET}")
    print(f"{BOLD}  BATTERY B — Consistency: C07 (06+09) × 3 runs{RESET}")
    print(f"  Tests that the address-rule fix is stable across multiple LLM calls.{RESET}")
    print(f"{BOLD}{'█'*72}{RESET}")
    results = []
    fnames = [FRONTEND_DEFAULTS[i]["name"] for i in range(len(FRONTEND_DEFAULTS))]
    for run_num in range(1, 4):
        print(f"\n{BOLD}  B-Run {run_num}/3{RESET}")
        parseds = []
        for k in ["06", "09"]:
            p = await _parse_ocr(FILES_0609[k])
            parseds.append(p)
        combined = combine_parsed_documents(parseds)
        usage = LLMUsage()
        t0 = time.perf_counter()
        rows = await extract_from_document(combined, FRONTEND_DEFAULTS, usage, force_sov=True)
        elapsed = time.perf_counter() - t0
        real = [r for r in rows if not r.get("_error")]
        fr = fill_rate(real, fnames)
        garble = find_garble(real, fnames)
        dups = find_dups(real)
        passed = len(real) == 20 and fr >= 0.50 and not garble and not dups
        print(f"  {_ok(len(real)==20, f'rows={len(real)}/20')}")
        print(f"  {_ok(fr>=0.50, f'fill={fr:.1%}')}")
        print(f"  {_ok(not garble, 'no garbled addresses')}")
        print(f"  {_ok(not dups, 'no duplicates')}")
        if garble:
            print(f"  {RED}GARBLE: {garble[:3]}{RESET}")
        # Check specifically the previously-garbled locs 5, 6, 16
        for r in real:
            if str(r.get("Loc #","")).strip() in {"5","6","16"}:
                print(f"    Loc {r['Loc #']}: Street={r.get('Street Address')!r} City={r.get('City')!r} TIV={r.get('Total Insurable Value (TIV)')!r}")
        verdict = f"{GREEN}PASS{RESET}" if passed else f"{RED}FAIL{RESET}"
        print(f"  Verdict: {verdict}  t={elapsed:.0f}s  ${usage.cost_usd:.4f}")
        results.append({"key": f"B-{run_num}", "pass": passed, "rows": len(real), "fill": fr, "garble": garble, "elapsed": elapsed, "cost": usage.cost_usd})
    return results


# ══════════════════════════════════════════════════════════════════════════════
# BATTERY C — Excel output inspection: C15 (all 4 docs)
# ══════════════════════════════════════════════════════════════════════════════

async def battery_c() -> List[Dict]:
    print(f"\n{BOLD}{'█'*72}{RESET}")
    print(f"{BOLD}  BATTERY C — Excel output inspection: C15 (06+07+08+09){RESET}")
    print(f"  Generates actual .xlsx and validates every cell.{RESET}")
    print(f"{BOLD}{'█'*72}{RESET}")

    parseds = []
    for k in ["06", "07", "08", "09"]:
        print(f"  OCR: {FILES_0609[k].name}")
        p = await _parse_ocr(FILES_0609[k])
        parseds.append(p)
    combined = combine_parsed_documents(parseds)

    usage = LLMUsage()
    t0 = time.perf_counter()
    rows = await extract_from_document(combined, FRONTEND_DEFAULTS, usage, force_sov=True)
    elapsed = time.perf_counter() - t0

    real = [r for r in rows if not r.get("_error")]
    fnames = [f["name"] for f in FRONTEND_DEFAULTS]
    fr = fill_rate(real, fnames)
    print(f"\n  Extraction: rows={len(real)}/20  fill={fr:.1%}  t={elapsed:.0f}s  ${usage.cost_usd:.4f}")

    # Generate Excel
    xlsx_bytes = generate_excel_bytes(real, fnames)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    data_rows = ws.max_row - 1
    print(f"  Excel: {data_rows} data rows  {ws.max_column} columns")
    print(f"  Headers: {headers[:8]}...")

    # Save for manual inspection
    out_path = _BACKEND / "test_files" / "C15_production_check.xlsx"
    out_path.write_bytes(xlsx_bytes)
    print(f"  Saved → {out_path.name}")

    issues: List[str] = []

    # Core columns that MUST be filled for every row
    MUST_FILL = ["Loc #", "Street Address", "City", "State", "Construction Type",
                 "Building Values", "Total Insurable Value (TIV)", "Year Built"]
    must_col = {fn: next((c for c in range(1, ws.max_column+1) if ws.cell(1,c).value == fn), None) for fn in MUST_FILL}

    print(f"\n  {BOLD}Cell-by-cell audit of core columns:{RESET}")
    for fn in MUST_FILL:
        col = must_col.get(fn)
        if col is None:
            issues.append(f"Column missing: {fn}")
            print(f"  {RED}✗ COLUMN MISSING: {fn}{RESET}")
            continue
        nulls = [r for r in range(2, ws.max_row+1) if not ws.cell(r, col).value]
        if nulls:
            issues.append(f"{fn}: null in rows {nulls}")
            print(f"  {RED}✗ {fn}: null in {len(nulls)} rows → rows {nulls[:5]}{RESET}")
        else:
            # Also show actual values
            vals = [ws.cell(r, col).value for r in range(2, min(ws.max_row+1, 6))]
            print(f"  {GREEN}✓ {fn}: all filled  e.g. {vals}{RESET}")

    # Value sanity checks on all rows
    print(f"\n  {BOLD}Value sanity checks:{RESET}")
    tiv_col  = must_col.get("Total Insurable Value (TIV)")
    yr_col   = must_col.get("Year Built")
    st_col   = must_col.get("State")
    loc_col  = must_col.get("Loc #")
    addr_col = must_col.get("Street Address")

    # TIV: must be numeric > 0
    tiv_issues = []
    if tiv_col:
        for r in range(2, ws.max_row+1):
            raw = ws.cell(r, tiv_col).value
            if raw:
                num_str = re.sub(r"[,$\s]", "", str(raw))
                try:
                    if float(num_str) <= 0:
                        tiv_issues.append(f"row {r}: {raw!r}")
                except ValueError:
                    tiv_issues.append(f"row {r}: non-numeric {raw!r}")
    print(f"  {_ok(not tiv_issues, f'TIV > 0 for all rows')}")
    if tiv_issues:
        for i in tiv_issues[:3]: print(f"    {RED}{i}{RESET}")

    # Year Built: 1870-2030
    yr_issues = []
    if yr_col:
        for r in range(2, ws.max_row+1):
            raw = ws.cell(r, yr_col).value
            if raw:
                try:
                    yr = int(str(raw).strip()[:4])
                    if not (1870 <= yr <= 2030):
                        yr_issues.append(f"row {r}: {raw!r}")
                except (ValueError, TypeError):
                    yr_issues.append(f"row {r}: bad format {raw!r}")
    print(f"  {_ok(not yr_issues, 'Year Built in range 1870-2030')}")
    if yr_issues:
        for i in yr_issues[:3]: print(f"    {RED}{i}{RESET}")

    # State: 2-char alpha
    st_issues = []
    if st_col:
        for r in range(2, ws.max_row+1):
            raw = str(ws.cell(r, st_col).value or "").strip()
            if raw and (len(raw) != 2 or not raw.isalpha()):
                st_issues.append(f"row {r}: {raw!r}")
    print(f"  {_ok(not st_issues, 'State is 2-char alpha code')}")
    if st_issues:
        for i in st_issues[:3]: print(f"    {RED}{i}{RESET}")

    # Loc #: no duplicates, sequential 1-20
    loc_vals = []
    if loc_col:
        loc_vals = [str(ws.cell(r, loc_col).value or "").strip() for r in range(2, ws.max_row+1)]
    dup_locs = [l for l in set(loc_vals) if loc_vals.count(l) > 1]
    print(f"  {_ok(not dup_locs, 'No duplicate Loc # values')}")
    if dup_locs:
        print(f"    {RED}Dups: {dup_locs}{RESET}")
    expected_locs = {str(i) for i in range(1, 21)}
    missing_locs = expected_locs - set(loc_vals)
    print(f"  {_ok(not missing_locs, 'All 20 locations present (1-20)')}")
    if missing_locs:
        print(f"    {RED}Missing: {sorted(missing_locs, key=int)}{RESET}")

    # Address garble check
    addr_garble = []
    if addr_col:
        for r in range(2, ws.max_row+1):
            val = str(ws.cell(r, addr_col).value or "").lower()
            for pat in _BAD_OCR:
                if pat in val:
                    addr_garble.append(f"row {r}: {ws.cell(r, addr_col).value!r}")
    print(f"  {_ok(not addr_garble, 'No garbled text in Street Address')}")

    # Print all 20 rows address+TIV table
    print(f"\n  {BOLD}All 20 rows — Loc / Street Address / City / State / TIV:{RESET}")
    for r in range(2, ws.max_row+1):
        loc  = ws.cell(r, loc_col).value  if loc_col  else "-"
        addr = ws.cell(r, addr_col).value if addr_col else "-"
        city = ws.cell(r, next((c for c in range(1,ws.max_column+1) if ws.cell(1,c).value=="City"), 0)).value if ws.max_column else "-"
        state= ws.cell(r, st_col).value   if st_col   else "-"
        tiv  = ws.cell(r, tiv_col).value  if tiv_col  else "-"
        print(f"    Loc {str(loc):>2}  {str(addr):<28}  {str(city):<16}  {str(state):<3}  {tiv}")

    passed = len(real) == 20 and not issues and not tiv_issues and not yr_issues and not st_issues and not dup_locs and not addr_garble
    verdict = f"{GREEN}PASS{RESET}" if passed else f"{RED}FAIL{RESET}"
    print(f"\n  Verdict: {verdict}")
    return [{"key": "C15-excel", "pass": passed, "rows": len(real), "fill": fr, "elapsed": elapsed, "cost": usage.cost_usd}]


# ══════════════════════════════════════════════════════════════════════════════
# BATTERY D — address quality across key combinations
# ══════════════════════════════════════════════════════════════════════════════

async def battery_d() -> List[Dict]:
    print(f"\n{BOLD}{'█'*72}{RESET}")
    print(f"{BOLD}  BATTERY D — Address quality: C01, C07, C11, C15{RESET}")
    print(f"  Inspects every Street Address, City, State, Zip for truncation/garble.{RESET}")
    print(f"{BOLD}{'█'*72}{RESET}")

    COMBOS = [
        ("C01", ["06"]),
        ("C07", ["06", "09"]),
        ("C11", ["06", "07", "08"]),
        ("C15", ["06", "07", "08", "09"]),
    ]
    ADDR_FIELDS = ["Street Address", "City", "State", "Zip"]
    results = []
    fnames = [f["name"] for f in FRONTEND_DEFAULTS]

    for label, keys in COMBOS:
        print(f"\n{BOLD}  D-{label}  [{'+'.join(keys)}]{RESET}")
        if len(keys) == 1:
            parsed = await asyncio.to_thread(parse_pdf, str(FILES_0609[keys[0]]), FILES_0609[keys[0]].name)
            usage = LLMUsage()
            t0 = time.perf_counter()
            rows = await extract_from_document(parsed, FRONTEND_DEFAULTS, usage, force_sov=True)
        else:
            parseds = []
            for k in keys:
                p = await _parse_ocr(FILES_0609[k])
                parseds.append(p)
            combined = combine_parsed_documents(parseds)
            usage = LLMUsage()
            t0 = time.perf_counter()
            rows = await extract_from_document(combined, FRONTEND_DEFAULTS, usage, force_sov=True)

        elapsed = time.perf_counter() - t0
        real = [r for r in rows if not r.get("_error")]
        garble = find_garble(real, fnames)
        dups = find_dups(real)

        # Address completeness
        addr_nulls: Dict[str, int] = {fn: 0 for fn in ADDR_FIELDS}
        addr_truncated: List[str] = []
        for r in real:
            for fn in ADDR_FIELDS:
                if not is_filled(r.get(fn)):
                    addr_nulls[fn] += 1
            addr = str(r.get("Street Address") or "")
            # Flag short addresses (likely truncated) — single word or very short
            if addr and len(addr.split()) <= 1 and len(addr) < 8:
                addr_truncated.append(f"Loc#{r.get('Loc #')}: {addr!r}")
            state = str(r.get("State") or "")
            if state and (len(state) != 2 or not state.isalpha()):
                addr_truncated.append(f"Loc#{r.get('Loc #')}: bad State={state!r}")

        print(f"  rows={len(real)}/20  t={elapsed:.0f}s  ${usage.cost_usd:.4f}")
        for fn in ADDR_FIELDS:
            null_ct = addr_nulls[fn]
            color = GREEN if null_ct == 0 else (YELLOW if null_ct <= 2 else RED)
            print(f"    {color}{fn}: {20-null_ct}/20 filled{RESET}", end="")
            print()
        if garble:
            print(f"  {RED}GARBLE: {garble}{RESET}")
        if addr_truncated:
            print(f"  {YELLOW}POSSIBLE TRUNCATION:{RESET}")
            for a in addr_truncated[:5]:
                print(f"    {YELLOW}{a}{RESET}")
        print(f"\n  {BOLD}All 20 addresses:{RESET}")
        for r in sorted(real, key=lambda x: int(str(x.get("Loc #","0")).strip() or "0") if str(x.get("Loc #","0")).strip().isdigit() else 99):
            street = r.get("Street Address") or "NULL"
            city   = r.get("City") or "NULL"
            state  = r.get("State") or "NULL"
            zipc   = r.get("Zip") or "NULL"
            tiv    = r.get("Total Insurable Value (TIV)") or "NULL"
            g_mark = f" {RED}←GARBLE{RESET}" if any(p in str(street).lower() for p in _BAD_OCR) else ""
            print(f"    Loc {str(r.get('Loc #','')):>2}: {street:<26} | {city:<16} | {state:<3} | {zipc:<10} | TIV={tiv}{g_mark}")

        passed = len(real) == 20 and not garble and not dups and not addr_truncated and all(v == 0 for v in addr_nulls.values())
        verdict = f"{GREEN}PASS{RESET}" if passed else (f"{YELLOW}WARN{RESET}" if len(real)==20 and not garble else f"{RED}FAIL{RESET}")
        print(f"  Verdict: {verdict}")
        results.append({"key": f"D-{label}", "pass": passed, "rows": len(real), "garble": garble, "elapsed": elapsed, "cost": usage.cost_usd})
    return results


# ══════════════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════════════

async def main() -> None:
    print(f"\n{BOLD}{'█'*72}{RESET}")
    print(f"{BOLD}  SOV PRODUCTION READINESS TEST{RESET}")
    print(f"  A: original files 01-05   B: C07 × 3 consistency")
    print(f"  C: Excel deep inspect     D: address quality × 4 combos")
    print(f"{BOLD}{'█'*72}{RESET}")

    all_results: List[Dict] = []

    ra = await battery_a()
    all_results.extend(ra)

    rb = await battery_b()
    all_results.extend(rb)

    rc = await battery_c()
    all_results.extend(rc)

    rd = await battery_d()
    all_results.extend(rd)

    # ── Grand summary ──────────────────────────────────────────────────────────
    print(f"\n\n{BOLD}{'█'*72}{RESET}")
    print(f"{BOLD}  GRAND SUMMARY{RESET}")
    print(f"{BOLD}{'█'*72}{RESET}")

    passed = [r for r in all_results if r.get("pass")]
    failed = [r for r in all_results if not r.get("pass")]
    total_cost = sum(r.get("cost", 0) for r in all_results)
    total_time = sum(r.get("elapsed", 0) for r in all_results)

    for r in all_results:
        sym = f"{GREEN}PASS{RESET}" if r.get("pass") else f"{RED}FAIL{RESET}"
        rows_str = f"{r.get('rows','?')}"
        exp_str  = f"/{r.get('expected','?')}" if r.get("expected") else ""
        fill_str = f"fill={r.get('fill',0):.1%}" if r.get("fill") is not None else ""
        garb_str = f" GARB:{len(r.get('garble',[]))}" if r.get("garble") else ""
        print(f"  {sym}  {r['key']:<12}  rows={rows_str+exp_str:<8}  {fill_str:<10}  {r.get('elapsed',0):.0f}s  ${r.get('cost',0):.3f}{garb_str}")

    print(f"\n  {BOLD}{len(passed)}/{len(all_results)} passed{RESET}  total_time={total_time:.0f}s  total_cost=${total_cost:.4f}")
    if failed:
        print(f"\n  {RED}FAILED:{RESET}")
        for r in failed:
            print(f"    {r['key']}: rows={r.get('rows')}, fill={r.get('fill',0):.1%}, garble={r.get('garble','')}")
    print()


if __name__ == "__main__":
    asyncio.run(main())
