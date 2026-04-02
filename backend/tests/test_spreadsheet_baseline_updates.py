from __future__ import annotations

import csv
import io
import sys
import unittest
from pathlib import Path

import openpyxl

_BACKEND = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_BACKEND))

from app.services.spreadsheet_service import update_csv_baseline_bytes, update_excel_baseline_bytes


class SpreadsheetBaselineUpdateTests(unittest.TestCase):
    def test_excel_updates_matched_rows_and_flags_missing_rows(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Location Number", "Street Address", "City", "State", "Zip", "Building Values", "Owner Note"])
        ws.append(["100", "100 Main St", "Austin", "TX", "78701", "100000", "keep"])
        ws.append(["200", "200 Main St", "Austin", "TX", "78702", "200000", "keep"])

        buf = io.BytesIO()
        wb.save(buf)

        output = update_excel_baseline_bytes(
            buf.getvalue(),
            [
                {
                    "Location Number": "100",
                    "Street Address": "100 Main St",
                    "City": "Austin",
                    "State": "TX",
                    "Zip": "78701",
                    "Building Values": "150000",
                    "_source_file": "renewal_a.pdf",
                },
                {
                    "Location Number": "300",
                    "Street Address": "300 Main St",
                    "City": "Austin",
                    "State": "TX",
                    "Zip": "78703",
                    "Building Values": "300000",
                    "_source_file": "renewal_b.pdf",
                },
            ],
            ["Location Number", "Street Address", "City", "State", "Zip", "Building Values"],
            True,
        )

        out_wb = openpyxl.load_workbook(io.BytesIO(output))
        out_ws = out_wb.active
        headers = [cell for cell in next(out_ws.iter_rows(min_row=1, max_row=1, values_only=True)) if cell]
        rows = list(out_ws.iter_rows(min_row=2, values_only=True))
        by_location = {str(row[0]): row for row in rows if row[0] is not None}
        status_idx = headers.index("GridPull Status")
        building_idx = headers.index("Building Values")
        owner_note_idx = headers.index("Owner Note")

        self.assertEqual(by_location["100"][building_idx], "150000")
        self.assertEqual(by_location["100"][owner_note_idx], "keep")
        self.assertEqual(by_location["100"][status_idx], "updated")

        self.assertEqual(by_location["200"][building_idx], "200000")
        self.assertEqual(by_location["200"][status_idx], "not_found")

        self.assertEqual(by_location["300"][building_idx], "300000")
        self.assertEqual(by_location["300"][status_idx], "new")

    def test_csv_preserves_matched_rows_when_overwrite_disabled(self) -> None:
        buf = io.StringIO()
        writer = csv.DictWriter(
            buf,
            fieldnames=["Street Address", "City", "State", "Zip", "Building Values", "Owner Note"],
        )
        writer.writeheader()
        writer.writerow(
            {
                "Street Address": "10 Main St",
                "City": "Austin",
                "State": "TX",
                "Zip": "78701",
                "Building Values": "100000",
                "Owner Note": "keep",
            }
        )
        writer.writerow(
            {
                "Street Address": "20 Main St",
                "City": "Austin",
                "State": "TX",
                "Zip": "78702",
                "Building Values": "200000",
                "Owner Note": "keep",
            }
        )

        output = update_csv_baseline_bytes(
            buf.getvalue().encode("utf-8"),
            [
                {
                    "Street Address": "10 Main St",
                    "City": "Austin",
                    "State": "TX",
                    "Zip": "78701",
                    "Building Values": "999999",
                    "_source_file": "renewal_c.pdf",
                },
                {
                    "Street Address": "30 Main St",
                    "City": "Austin",
                    "State": "TX",
                    "Zip": "78703",
                    "Building Values": "300000",
                    "_source_file": "renewal_d.pdf",
                },
            ],
            ["Street Address", "City", "State", "Zip", "Building Values"],
            False,
        )

        reader = csv.DictReader(io.StringIO(output.decode("utf-8")))
        rows = list(reader)
        by_address = {row["Street Address"]: row for row in rows}

        self.assertEqual(by_address["10 Main St"]["Building Values"], "100000")
        self.assertEqual(by_address["10 Main St"]["Owner Note"], "keep")
        self.assertEqual(by_address["10 Main St"]["GridPull Status"], "matched_preserved")

        self.assertEqual(by_address["20 Main St"]["Building Values"], "200000")
        self.assertEqual(by_address["20 Main St"]["GridPull Status"], "not_found")

        self.assertEqual(by_address["30 Main St"]["Building Values"], "300000")
        self.assertEqual(by_address["30 Main St"]["GridPull Status"], "new")


    def test_excel_matches_vehicle_rows_by_vehicle_number(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Vehicle #", "Year", "Make", "Model", "VIN", "Cost New"])
        ws.append(["1", "2023", "Toyota", "Tacoma", "VIN001", "35000"])
        ws.append(["2", "2022", "Ford", "F-150", "VIN002", "42000"])

        buf = io.BytesIO()
        wb.save(buf)

        output = update_excel_baseline_bytes(
            buf.getvalue(),
            [
                {
                    "Vehicle #": "1",
                    "Year": "2023",
                    "Make": "Toyota",
                    "Model": "Tacoma",
                    "VIN": "VIN001",
                    "Cost New": "36000",
                    "_source_file": "renewal.pdf",
                },
                {
                    "Vehicle #": "3",
                    "Year": "2024",
                    "Make": "Chevy",
                    "Model": "Silverado",
                    "VIN": "VIN003",
                    "Cost New": "50000",
                    "_source_file": "renewal.pdf",
                },
            ],
            ["Vehicle #", "Year", "Make", "Model", "VIN", "Cost New"],
            True,
        )

        out_wb = openpyxl.load_workbook(io.BytesIO(output))
        out_ws = out_wb.active
        headers = [cell for cell in next(out_ws.iter_rows(min_row=1, max_row=1, values_only=True)) if cell]
        rows = list(out_ws.iter_rows(min_row=2, values_only=True))
        by_veh = {str(row[0]): row for row in rows if row[0] is not None}
        status_idx = headers.index("GridPull Status")
        cost_idx = headers.index("Cost New")

        self.assertEqual(by_veh["1"][cost_idx], "36000")
        self.assertEqual(by_veh["1"][status_idx], "updated")
        self.assertEqual(by_veh["2"][status_idx], "not_found")
        self.assertEqual(by_veh["3"][cost_idx], "50000")
        self.assertEqual(by_veh["3"][status_idx], "new")

    def test_excel_matches_by_address_when_no_location_field(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Street Address", "City", "State", "Zip", "Building Values"])
        ws.append(["100 Main St", "Austin", "TX", "78701", "100000"])
        ws.append(["200 Oak Ave", "Dallas", "TX", "75201", "200000"])

        buf = io.BytesIO()
        wb.save(buf)

        output = update_excel_baseline_bytes(
            buf.getvalue(),
            [
                {"Street Address": "100 Main St", "City": "Austin", "State": "TX", "Zip": "78701", "Building Values": "150000", "_source_file": "renewal.pdf"},
                {"Street Address": "300 Elm Rd", "City": "Houston", "State": "TX", "Zip": "77001", "Building Values": "300000", "_source_file": "renewal.pdf"},
            ],
            ["Street Address", "City", "State", "Zip", "Building Values"],
            True,
        )

        out_wb = openpyxl.load_workbook(io.BytesIO(output))
        out_ws = out_wb.active
        headers = [cell for cell in next(out_ws.iter_rows(min_row=1, max_row=1, values_only=True)) if cell]
        rows = list(out_ws.iter_rows(min_row=2, values_only=True))
        status_idx = headers.index("GridPull Status")
        addr_idx = headers.index("Street Address")
        bldg_idx = headers.index("Building Values")

        by_addr = {str(r[addr_idx]): r for r in rows if r[addr_idx]}
        self.assertEqual(by_addr["100 Main St"][bldg_idx], "150000")
        self.assertEqual(by_addr["100 Main St"][status_idx], "updated")
        self.assertEqual(by_addr["200 Oak Ave"][status_idx], "not_found")
        self.assertEqual(by_addr["300 Elm Rd"][bldg_idx], "300000")
        self.assertEqual(by_addr["300 Elm Rd"][status_idx], "new")

    def test_excel_handles_duplicate_location_numbers_gracefully(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Location Number", "Address", "City", "State", "Zip", "TIV"])
        ws.append(["1", "100 Main St", "Austin", "TX", "78701", "100000"])
        ws.append(["1", "200 Oak Ave", "Austin", "TX", "78702", "200000"])

        buf = io.BytesIO()
        wb.save(buf)

        output = update_excel_baseline_bytes(
            buf.getvalue(),
            [
                {"Location Number": "1", "Address": "100 Main St", "City": "Austin", "State": "TX", "Zip": "78701", "TIV": "150000", "_source_file": "renewal.pdf"},
            ],
            ["Location Number", "Address", "City", "State", "Zip", "TIV"],
            True,
        )

        out_wb = openpyxl.load_workbook(io.BytesIO(output))
        out_ws = out_wb.active
        rows = list(out_ws.iter_rows(min_row=2, values_only=True))
        non_empty = [r for r in rows if any(c is not None and str(c).strip() for c in r)]
        self.assertEqual(len(non_empty), 3)

    def test_excel_matches_by_invoice_number(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Invoice #", "Date", "Amount", "Vendor"])
        ws.append(["INV-001", "2024-01-15", "5000", "Acme"])
        ws.append(["INV-002", "2024-02-20", "3000", "Globex"])

        buf = io.BytesIO()
        wb.save(buf)

        output = update_excel_baseline_bytes(
            buf.getvalue(),
            [
                {"Invoice #": "INV-001", "Date": "2024-01-15", "Amount": "5500", "Vendor": "Acme", "_source_file": "q2.pdf"},
                {"Invoice #": "INV-003", "Date": "2024-03-10", "Amount": "7000", "Vendor": "Initech", "_source_file": "q2.pdf"},
            ],
            ["Invoice #", "Date", "Amount", "Vendor"],
            True,
        )

        out_wb = openpyxl.load_workbook(io.BytesIO(output))
        out_ws = out_wb.active
        headers = [cell for cell in next(out_ws.iter_rows(min_row=1, max_row=1, values_only=True)) if cell]
        rows = list(out_ws.iter_rows(min_row=2, values_only=True))
        by_inv = {str(row[0]): row for row in rows if row[0]}
        status_idx = headers.index("GridPull Status")
        amount_idx = headers.index("Amount")

        self.assertEqual(by_inv["INV-001"][amount_idx], "5500")
        self.assertEqual(by_inv["INV-001"][status_idx], "updated")
        self.assertEqual(by_inv["INV-002"][status_idx], "not_found")
        self.assertEqual(by_inv["INV-003"][status_idx], "new")

    def test_phone_number_field_is_not_treated_as_id(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Phone #", "City", "State", "Zip"])
        ws.append(["Alice", "555-0001", "Austin", "TX", "78701"])

        buf = io.BytesIO()
        wb.save(buf)

        output = update_excel_baseline_bytes(
            buf.getvalue(),
            [
                {"Name": "Bob", "Phone #": "555-0002", "City": "Dallas", "State": "TX", "Zip": "75201", "_source_file": "t.pdf"},
            ],
            ["Name", "Phone #", "City", "State", "Zip"],
            True,
        )

        out_wb = openpyxl.load_workbook(io.BytesIO(output))
        out_ws = out_wb.active
        headers = [cell for cell in next(out_ws.iter_rows(min_row=1, max_row=1, values_only=True)) if cell]
        rows = list(out_ws.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len([r for r in rows if any(c is not None and str(c).strip() for c in r)]), 2)

    def test_csv_matches_employee_rows_by_employee_id(self) -> None:
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=["Employee #", "Name", "Department", "Salary"])
        writer.writeheader()
        writer.writerow({"Employee #": "E001", "Name": "Alice", "Department": "Eng", "Salary": "90000"})
        writer.writerow({"Employee #": "E002", "Name": "Bob", "Department": "Sales", "Salary": "85000"})

        output = update_csv_baseline_bytes(
            buf.getvalue().encode("utf-8"),
            [
                {"Employee #": "E001", "Name": "Alice", "Department": "Eng", "Salary": "95000", "_source_file": "q2.pdf"},
                {"Employee #": "E003", "Name": "Carol", "Department": "HR", "Salary": "88000", "_source_file": "q2.pdf"},
            ],
            ["Employee #", "Name", "Department", "Salary"],
            True,
        )

        reader = csv.DictReader(io.StringIO(output.decode("utf-8")))
        rows = list(reader)
        by_emp = {row["Employee #"]: row for row in rows}

        self.assertEqual(by_emp["E001"]["Salary"], "95000")
        self.assertEqual(by_emp["E001"]["GridPull Status"], "updated")
        self.assertEqual(by_emp["E002"]["GridPull Status"], "not_found")
        self.assertEqual(by_emp["E003"]["Salary"], "88000")
        self.assertEqual(by_emp["E003"]["GridPull Status"], "new")


if __name__ == "__main__":
    unittest.main()
