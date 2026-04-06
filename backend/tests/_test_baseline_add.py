import asyncio, io, sys
from pathlib import Path
sys.path.insert(0, str(Path(".").resolve()))
import os
try:
    from dotenv import load_dotenv; load_dotenv(".env")
except ImportError: pass
os.environ.pop("STRIPE_PRODUCT_ID", None)

import openpyxl
from app.services.extraction import extract_from_document, LLMUsage
from app.services.pdf_service import parse_pdf
from app.services.spreadsheet_service import update_excel_baseline_bytes
from tests.test_sov_e2e import FIELD_SETS

async def main():
    fields = FIELD_SETS["frontend_defaults"]
    field_names = [f["name"] for f in fields]

    pdf_path = Path("test_files/01_property_appraisal_report_25_buildings.pdf")
    baseline_path = Path("test_files/01_baseline_25_buildings_updated.xlsx")

    print(f"Baseline: {baseline_path.name}")
    baseline_bytes = baseline_path.read_bytes()
    wb0 = openpyxl.load_workbook(io.BytesIO(baseline_bytes))
    ws0 = wb0.active
    print(f"  Existing rows: {ws0.max_row - 1}")
    status_col = next((c for c in range(1, ws0.max_column+1) if ws0.cell(1,c).value == "GridPull Status"), None)
    if status_col:
        statuses = {}
        for r in range(2, ws0.max_row+1):
            v = ws0.cell(r, status_col).value or "none"
            statuses[v] = statuses.get(v, 0) + 1
        print(f"  Existing statuses: {statuses}")

    print(f"\nExtracting from {pdf_path.name}...")
    parsed = parse_pdf(str(pdf_path), pdf_path.name)
    usage = LLMUsage()
    rows = await extract_from_document(parsed, fields, usage)
    real_rows = [r for r in rows if not r.get("_error")]
    print(f"  Extracted {len(real_rows)} rows  cost=${usage.cost_usd:.4f}")

    print("\nUpdating baseline (allow_edit=False — preserve existing values, fill gaps)...")
    updated = update_excel_baseline_bytes(baseline_bytes, real_rows, field_names, allow_edit_past_values=False)

    wb = openpyxl.load_workbook(io.BytesIO(updated))
    ws = wb.active
    data_rows = ws.max_row - 1
    status_col2 = next((c for c in range(1, ws.max_column+1) if ws.cell(1,c).value == "GridPull Status"), None)
    statuses2 = {}
    if status_col2:
        for r in range(2, ws.max_row+1):
            v = ws.cell(r, status_col2).value or "none"
            statuses2[v] = statuses2.get(v, 0) + 1

    print(f"  Output rows: {data_rows}")
    print(f"  Status distribution: {statuses2}")

    tiv_col = next((c for c in range(1, ws.max_column+1) if ws.cell(1,c).value == "Total Insurable Value (TIV)"), None)
    loc_col = next((c for c in range(1, ws.max_column+1) if ws.cell(1,c).value == "Location Name"), None)
    tiv_filled = sum(1 for r in range(2, ws.max_row+1) if tiv_col and ws.cell(r, tiv_col).value)
    loc_filled = sum(1 for r in range(2, ws.max_row+1) if loc_col and ws.cell(r, loc_col).value)
    print(f"  TIV filled: {tiv_filled}/{data_rows}")
    print(f"  Location Name filled: {loc_filled}/{data_rows}")

    print("\nSample rows (first 3):")
    for r in range(2, min(5, ws.max_row+1)):
        vals = {ws.cell(1,c).value: ws.cell(r,c).value for c in range(2, 9)}
        print(f"  Row {r-1}: {vals}")

    out = Path("test_files/01_baseline_add_test_result.xlsx")
    out.write_bytes(updated)
    print(f"\nSaved result -> {out.name}")

asyncio.run(main())
