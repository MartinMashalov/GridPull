"""
End-to-end SOV testing: creation from scratch + baseline spreadsheet updates.

Tests both scenarios:
  1. Extract from PDFs → create new Excel/CSV from scratch
  2. Extract from PDFs → update an existing baseline spreadsheet (with/without edit)

Also tests diverse field-name variations to ensure robustness.

Usage:
    cd backend
    python tests/test_sov_e2e.py [--files 01 02 03 04 05] [--skip-extraction]
"""
from __future__ import annotations

import asyncio
import csv
import io
import json
import sys
import time
from pathlib import Path
from typing import Any, Dict, List

import openpyxl

_BACKEND = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_BACKEND))

import os
try:
    from dotenv import load_dotenv
    load_dotenv(_BACKEND / ".env")
except ImportError:
    pass
os.environ.pop("STRIPE_PRODUCT_ID", None)

from app.services.extraction import extract_from_document, LLMUsage
from app.services.pdf_service import parse_pdf
from app.services.spreadsheet_service import (
    generate_excel_bytes,
    generate_csv_bytes,
    update_excel_baseline_bytes,
    update_csv_baseline_bytes,
    read_headers_from_bytes,
)

GREEN = "\033[92m"
YELLOW = "\033[93m"
RED = "\033[91m"
BOLD = "\033[1m"
DIM = "\033[2m"
RESET = "\033[0m"
CYAN = "\033[96m"

_EMPTY_VALUES = {"", "null", "none", "n/a", "na", "-", "—", "unknown", "not found", "not available"}

TEST_FILES = {
    "01": (_BACKEND / "test_files" / "01_property_appraisal_report_25_buildings.pdf", 25),
    "02": (_BACKEND / "test_files" / "02_carrier_property_schedule_50_locations.pdf", 50),
    "03": (_BACKEND / "test_files" / "03_package_policy_dec_pages_15_locations.pdf", 15),
    "04": (_BACKEND / "test_files" / "04_vehicle_schedule_35_vehicles.pdf", 35),
    "05": (_BACKEND / "test_files" / "05_prior_year_sov_30_locations.pdf", 30),
}

# ──────────────────────────────────────────────────────────────────────────────
# Field-name variations to test robustness
# ──────────────────────────────────────────────────────────────────────────────

FIELD_SETS = {
    "frontend_defaults": [
        {"name": "Location Number", "description": "Extract the schedule location identifier exactly as shown."},
        {"name": "Address Line 1", "description": "Extract the primary street address for the insured location."},
        {"name": "City", "description": "Extract only the city name tied to the location address."},
        {"name": "State", "description": "Extract the state or province code, preferring postal abbreviation."},
        {"name": "ZIP Code", "description": "Extract the postal code for the location exactly as displayed."},
        {"name": "Construction Class", "description": "Extract the insurance construction classification."},
        {"name": "Year Built", "description": "Extract the original year of construction for the building."},
        {"name": "Building Value", "description": "Extract the insured building amount for the location."},
        {"name": "Contents / BPP Value", "description": "Extract the insured contents or business personal property amount."},
        {"name": "Business Income Value", "description": "Extract the business income or time-element insured amount."},
        {"name": "Total Insured Value", "description": "Extract the total insured value for the location."},
    ],
    "abbreviated": [
        {"name": "Loc #", "description": "Location or property number/ID"},
        {"name": "Bldg #", "description": "Building number"},
        {"name": "Street Address", "description": "Street address of the property"},
        {"name": "City", "description": "City"},
        {"name": "State", "description": "State (2-letter code)"},
        {"name": "Zip", "description": "Zip code"},
        {"name": "Const Type", "description": "Construction type (e.g. frame, masonry)"},
        {"name": "Yr Built", "description": "Year the building was constructed"},
        {"name": "Bldg Val", "description": "Building insured value in dollars"},
        {"name": "BPP Val", "description": "Contents or BPP value in dollars"},
        {"name": "TIV", "description": "Total insured value in dollars"},
    ],
    "verbose": [
        {"name": "Property Location Number", "description": "The unique identifier for each insured location/premises"},
        {"name": "Property Street Address", "description": "Full street address including suite or unit number"},
        {"name": "Property City", "description": "City where the insured property is located"},
        {"name": "Property State", "description": "State code (2-letter) for the insured property"},
        {"name": "Property Zip Code", "description": "5-digit or ZIP+4 postal code"},
        {"name": "Construction Type", "description": "Construction type/class used by underwriters"},
        {"name": "Occupancy Type", "description": "Occupancy or exposure classification"},
        {"name": "Year Built", "description": "Original year of construction"},
        {"name": "Building Replacement Cost", "description": "Building insured/replacement cost value"},
        {"name": "Personal Property Value", "description": "Contents or business personal property value"},
        {"name": "Total Insurable Value (TIV)", "description": "Sum of all insured value components"},
    ],
    "vehicle_focused": [
        {"name": "Vehicle #", "description": "Vehicle or unit number"},
        {"name": "Year", "description": "Vehicle model year"},
        {"name": "Make", "description": "Vehicle manufacturer"},
        {"name": "Model", "description": "Vehicle model name"},
        {"name": "VIN", "description": "Vehicle identification number"},
        {"name": "Body Type", "description": "Vehicle body type (e.g. Sedan, Pickup, Van)"},
        {"name": "Cost New", "description": "Original cost or stated value of the vehicle"},
        {"name": "Garaging City", "description": "City where the vehicle is garaged"},
        {"name": "Garaging State", "description": "State where the vehicle is garaged"},
        {"name": "Garaging Zip", "description": "Zip code where the vehicle is garaged"},
        {"name": "Radius", "description": "Operating radius in miles"},
    ],
    "mixed_case_symbols": [
        {"name": "LOC #", "description": "Location identifier"},
        {"name": "ADDRESS", "description": "Street address"},
        {"name": "CITY", "description": "City name"},
        {"name": "ST", "description": "State abbreviation"},
        {"name": "ZIP", "description": "Postal code"},
        {"name": "CONSTRUCTION", "description": "Construction type"},
        {"name": "YR BUILT", "description": "Year of construction"},
        {"name": "BLDG VALUE ($)", "description": "Building value in dollars"},
        {"name": "CONTENTS ($)", "description": "Contents value in dollars"},
        {"name": "BI/EE ($)", "description": "Business income / extra expense value"},
        {"name": "TIV ($)", "description": "Total insured value"},
    ],
}

FIELD_SETS["carrier_style"] = [
    {"name": "Loc No.", "description": "Location or premises number"},
    {"name": "Premises Address", "description": "Street address of the insured premises"},
    {"name": "City", "description": "City name"},
    {"name": "State", "description": "State (2-letter code)"},
    {"name": "ZIP", "description": "5-digit zip code"},
    {"name": "Building Limit", "description": "Building coverage limit in dollars"},
    {"name": "BPP Limit", "description": "Business personal property limit"},
    {"name": "BI Limit", "description": "Business income limit"},
    {"name": "Total Limit", "description": "Total combined coverage limit"},
]

FIELD_SETS["reinsurance_style"] = [
    {"name": "Site ID", "description": "Unique site identifier"},
    {"name": "Street", "description": "Street address"},
    {"name": "City", "description": "City"},
    {"name": "State", "description": "State or province"},
    {"name": "Postal Code", "description": "Zip or postal code"},
    {"name": "Constr Code", "description": "Construction code (ISO)"},
    {"name": "Occ Code", "description": "Occupancy code"},
    {"name": "Year Built", "description": "Year constructed"},
    {"name": "Bldg RCV", "description": "Building replacement cost value"},
    {"name": "Contents RCV", "description": "Contents replacement cost value"},
    {"name": "BI/EE", "description": "Business income / extra expense"},
    {"name": "TIV", "description": "Total insured value"},
]

FIELD_SETS["minimal"] = [
    {"name": "Location #", "description": "Location number or ID"},
    {"name": "Address", "description": "Full street address"},
    {"name": "City", "description": "City"},
    {"name": "State", "description": "State"},
    {"name": "Zip", "description": "Zip code"},
    {"name": "TIV", "description": "Total insured value"},
]

RECOMMENDED_FIELD_SET_PER_FILE = {
    "01": "frontend_defaults",
    "02": "abbreviated",
    "03": "verbose",
    "04": "vehicle_focused",
    "05": "mixed_case_symbols",
}


def is_filled(value: Any) -> bool:
    if value is None:
        return False
    return str(value).strip().lower() not in _EMPTY_VALUES


def compute_fill_rate(rows: List[Dict[str, Any]], field_names: List[str]) -> float:
    total = len(rows) * len(field_names)
    if total == 0:
        return 0
    filled = sum(1 for r in rows for fn in field_names if is_filled(r.get(fn)))
    return filled / total


def print_extraction_report(
    rows: List[Dict[str, Any]],
    field_names: List[str],
    expected_rows: int,
    label: str,
    elapsed: float,
    cost: float,
) -> dict:
    real_rows = [r for r in rows if not r.get("_error")]
    error_rows = [r for r in rows if r.get("_error")]
    fill_rate = compute_fill_rate(real_rows, field_names)

    print(f"\n  {CYAN}--- {label} ---{RESET}")
    print(f"  Rows: {BOLD}{len(real_rows)}{RESET} (expected ~{expected_rows})")
    if error_rows:
        print(f"  {RED}Error rows: {len(error_rows)}{RESET}")
    print(f"  Fill rate: {fill_rate:.1%}")
    print(f"  Time: {elapsed:.1f}s  Cost: ${cost:.6f}")

    print(f"  {DIM}Per-field:{RESET}")
    for fn in field_names:
        filled = sum(1 for r in real_rows if is_filled(r.get(fn)))
        pct = filled / len(real_rows) if real_rows else 0
        color = GREEN if pct >= 0.7 else (YELLOW if pct >= 0.3 else RED)
        bar = "█" * int(pct * 10) + "░" * (10 - int(pct * 10))
        print(f"    {color}{bar} {pct:5.0%} {fn}{RESET}")

    count_ok = abs(len(real_rows) - expected_rows) <= max(3, int(expected_rows * 0.15))
    fill_ok = fill_rate >= 0.30

    if count_ok and fill_ok:
        verdict = f"{GREEN}PASS{RESET}"
    elif count_ok or fill_ok:
        verdict = f"{YELLOW}PARTIAL{RESET}"
    else:
        verdict = f"{RED}FAIL{RESET}"
    print(f"  Verdict: {verdict}")

    return {
        "rows": len(real_rows),
        "expected": expected_rows,
        "fill_rate": fill_rate,
        "errors": len(error_rows),
        "count_ok": count_ok,
        "fill_ok": fill_ok,
        "pass": count_ok and fill_ok,
    }


async def test_create_from_scratch(
    file_key: str,
    field_set_name: str,
) -> dict | None:
    path, expected = TEST_FILES[file_key]
    if not path.exists():
        print(f"{RED}  File not found: {path}{RESET}")
        return None

    fields = FIELD_SETS[field_set_name]
    field_names = [f["name"] for f in fields]

    print(f"\n{BOLD}{'═' * 72}{RESET}")
    print(f"{BOLD}  SCENARIO 1: Create from Scratch{RESET}")
    print(f"  File: {path.name}")
    print(f"  Field set: {field_set_name} ({len(fields)} fields)")
    print(f"{BOLD}{'═' * 72}{RESET}")

    parsed = parse_pdf(str(path), path.name)
    print(f"  Parsed: {parsed.page_count} pages, {len(parsed.tables)} tables, scanned={parsed.is_scanned}")

    usage = LLMUsage()
    t0 = time.perf_counter()
    rows = await extract_from_document(parsed, fields, usage)
    elapsed = time.perf_counter() - t0

    report = print_extraction_report(rows, field_names, expected, "Extraction", elapsed, usage.cost_usd)

    real_rows = [r for r in rows if not r.get("_error")]
    if not real_rows:
        print(f"  {RED}No real rows extracted, skipping spreadsheet generation{RESET}")
        return report

    excel_bytes = generate_excel_bytes(real_rows, field_names)
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active
    excel_headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    excel_data_rows = ws.max_row - 1
    print(f"\n  {CYAN}Excel generation:{RESET}")
    print(f"    Headers: {excel_headers}")
    print(f"    Data rows: {excel_data_rows}")
    if excel_data_rows != len(real_rows):
        print(f"    {RED}MISMATCH: expected {len(real_rows)} data rows in Excel{RESET}")
        report["excel_ok"] = False
    else:
        print(f"    {GREEN}OK{RESET}")
        report["excel_ok"] = True

    csv_bytes = generate_csv_bytes(real_rows, field_names)
    reader = csv.DictReader(io.StringIO(csv_bytes.decode("utf-8")))
    csv_rows = list(reader)
    csv_headers = reader.fieldnames
    print(f"\n  {CYAN}CSV generation:{RESET}")
    print(f"    Headers: {csv_headers}")
    print(f"    Data rows: {len(csv_rows)}")
    if len(csv_rows) != len(real_rows):
        print(f"    {RED}MISMATCH: expected {len(real_rows)} data rows in CSV{RESET}")
        report["csv_ok"] = False
    else:
        print(f"    {GREEN}OK{RESET}")
        report["csv_ok"] = True

    print(f"\n  {BOLD}Sample rows (first 3):{RESET}")
    for i, row in enumerate(real_rows[:3]):
        vals = {fn: row.get(fn, "-") for fn in field_names[:6]}
        print(f"    Row {i+1}: {vals}")

    return report


async def test_baseline_update(
    file_key: str,
    field_set_name: str,
    extracted_rows: List[Dict[str, Any]] | None = None,
) -> dict | None:
    path, expected = TEST_FILES[file_key]
    if not path.exists():
        print(f"{RED}  File not found: {path}{RESET}")
        return None

    fields = FIELD_SETS[field_set_name]
    field_names = [f["name"] for f in fields]

    print(f"\n{BOLD}{'═' * 72}{RESET}")
    print(f"{BOLD}  SCENARIO 2: Update Existing Spreadsheet{RESET}")
    print(f"  File: {path.name}")
    print(f"  Field set: {field_set_name}")
    print(f"{BOLD}{'═' * 72}{RESET}")

    if extracted_rows is None:
        parsed = parse_pdf(str(path), path.name)
        usage = LLMUsage()
        t0 = time.perf_counter()
        rows = await extract_from_document(parsed, fields, usage)
        elapsed = time.perf_counter() - t0
        extracted_rows = [r for r in rows if not r.get("_error")]
        print(f"  Extracted {len(extracted_rows)} rows in {elapsed:.1f}s (${usage.cost_usd:.6f})")

    if not extracted_rows or len(extracted_rows) < 4:
        print(f"  {RED}Not enough rows to test baseline update{RESET}")
        return None

    split = len(extracted_rows) * 2 // 3
    first_batch = extracted_rows[:split]
    second_batch = extracted_rows[split:]
    overlap_rows = extracted_rows[split - 3 : split]

    # ── Sub-test A: Excel baseline with allow_edit=True ──
    print(f"\n  {CYAN}Sub-test A: Excel baseline, allow_edit=True{RESET}")
    print(f"    Creating baseline with {len(first_batch)} rows...")
    print(f"    Update batch: {len(second_batch)} new + {len(overlap_rows)} overlapping (modified)")

    initial_excel = generate_excel_bytes(first_batch, field_names)
    headers_read = read_headers_from_bytes(initial_excel, "xlsx")
    print(f"    Baseline headers: {headers_read[:5]}...")

    combined_update = list(second_batch)
    for overlap_row in overlap_rows:
        modified_row = dict(overlap_row)
        for fn in reversed(field_names):
            if fn.lower() not in ("location number", "loc #", "loc#", "vehicle #", "property location number", "address", "address line 1", "street address", "property street address", "city", "state", "zip", "zip code", "property city", "property state", "property zip code", "st", "lot #") and is_filled(modified_row.get(fn)):
                modified_row[fn] = str(modified_row[fn]) + " UPDATED"
                break
        combined_update.insert(0, modified_row)

    updated_excel = update_excel_baseline_bytes(
        initial_excel, combined_update, field_names, True
    )
    wb = openpyxl.load_workbook(io.BytesIO(updated_excel))
    ws = wb.active
    out_headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value]
    out_rows = list(ws.iter_rows(min_row=2, values_only=True))
    out_rows = [r for r in out_rows if any(c is not None and str(c).strip() for c in r)]

    status_col = None
    for i, h in enumerate(out_headers):
        if h == "GridPull Status":
            status_col = i
            break

    statuses = {}
    if status_col is not None:
        for r in out_rows:
            st = r[status_col] if status_col < len(r) else None
            if st:
                statuses[st] = statuses.get(st, 0) + 1

    print(f"    Output rows: {len(out_rows)}")
    print(f"    Status distribution: {statuses}")
    print(f"    Headers: {out_headers[:6]}...")

    excel_a_ok = len(out_rows) >= len(extracted_rows)
    if "updated" in statuses or "new" in statuses:
        print(f"    {GREEN}OK - found updated/new rows{RESET}")
    else:
        print(f"    {YELLOW}WARNING - no updated/new status found{RESET}")
        excel_a_ok = False

    # ── Sub-test B: Excel baseline with allow_edit=False ──
    print(f"\n  {CYAN}Sub-test B: Excel baseline, allow_edit=False (preserve){RESET}")
    updated_preserve = update_excel_baseline_bytes(
        initial_excel, combined_update, field_names, False
    )
    wb2 = openpyxl.load_workbook(io.BytesIO(updated_preserve))
    ws2 = wb2.active
    out_headers2 = [ws2.cell(row=1, column=c).value for c in range(1, ws2.max_column + 1) if ws2.cell(row=1, column=c).value]
    out_rows2 = list(ws2.iter_rows(min_row=2, values_only=True))
    out_rows2 = [r for r in out_rows2 if any(c is not None and str(c).strip() for c in r)]

    statuses2 = {}
    if status_col is not None:
        for idx, h in enumerate(out_headers2):
            if h == "GridPull Status":
                status_col2 = idx
                break
        else:
            status_col2 = None
        if status_col2 is not None:
            for r in out_rows2:
                st = r[status_col2] if status_col2 < len(r) else None
                if st:
                    statuses2[st] = statuses2.get(st, 0) + 1

    print(f"    Output rows: {len(out_rows2)}")
    print(f"    Status distribution: {statuses2}")

    excel_b_ok = True
    if "matched_preserved" in statuses2:
        print(f"    {GREEN}OK - matched rows preserved{RESET}")
    elif "not_found" in statuses2 or "new" in statuses2:
        print(f"    {GREEN}OK - found not_found/new statuses{RESET}")
    else:
        print(f"    {YELLOW}WARNING - no matched_preserved status{RESET}")

    # ── Sub-test C: CSV baseline with allow_edit=True ──
    print(f"\n  {CYAN}Sub-test C: CSV baseline, allow_edit=True{RESET}")
    initial_csv = generate_csv_bytes(first_batch, field_names)
    updated_csv = update_csv_baseline_bytes(
        initial_csv, combined_update, field_names, True
    )
    reader = csv.DictReader(io.StringIO(updated_csv.decode("utf-8")))
    csv_out_rows = list(reader)
    csv_statuses = {}
    for r in csv_out_rows:
        st = r.get("GridPull Status")
        if st:
            csv_statuses[st] = csv_statuses.get(st, 0) + 1

    print(f"    Output rows: {len(csv_out_rows)}")
    print(f"    Status distribution: {csv_statuses}")

    csv_ok = len(csv_out_rows) >= len(extracted_rows)
    if "updated" in csv_statuses or "new" in csv_statuses:
        print(f"    {GREEN}OK{RESET}")
    else:
        print(f"    {YELLOW}WARNING{RESET}")
        csv_ok = False

    # ── Sub-test D: CSV baseline with allow_edit=False ──
    print(f"\n  {CYAN}Sub-test D: CSV baseline, allow_edit=False{RESET}")
    preserved_csv = update_csv_baseline_bytes(
        initial_csv, combined_update, field_names, False
    )
    reader2 = csv.DictReader(io.StringIO(preserved_csv.decode("utf-8")))
    csv_out_rows2 = list(reader2)
    csv_statuses2 = {}
    for r in csv_out_rows2:
        st = r.get("GridPull Status")
        if st:
            csv_statuses2[st] = csv_statuses2.get(st, 0) + 1

    print(f"    Output rows: {len(csv_out_rows2)}")
    print(f"    Status distribution: {csv_statuses2}")
    if "matched_preserved" in csv_statuses2:
        print(f"    {GREEN}OK - preserved matched rows{RESET}")
    else:
        print(f"    {YELLOW}WARNING{RESET}")

    return {
        "excel_edit_ok": excel_a_ok,
        "excel_preserve_ok": excel_b_ok,
        "csv_edit_ok": csv_ok,
        "total_rows_a": len(out_rows),
        "total_rows_b": len(out_rows2),
        "total_rows_c": len(csv_out_rows),
    }


async def test_multi_field_set_extraction(file_key: str) -> list:
    """Run extraction against the same file with all applicable field sets."""
    path, expected = TEST_FILES[file_key]
    if not path.exists():
        print(f"{RED}  File not found: {path}{RESET}")
        return []

    print(f"\n{BOLD}{'═' * 72}{RESET}")
    print(f"{BOLD}  ROBUSTNESS: Multiple field sets for {path.name}{RESET}")
    print(f"{BOLD}{'═' * 72}{RESET}")

    parsed = parse_pdf(str(path), path.name)
    print(f"  Parsed: {parsed.page_count} pages, {len(parsed.tables)} tables, scanned={parsed.is_scanned}")

    applicable = ["frontend_defaults", "abbreviated", "verbose", "mixed_case_symbols"]
    if file_key == "04":
        applicable = ["vehicle_focused", "abbreviated"]

    results = []
    for fs_name in applicable:
        fields = FIELD_SETS[fs_name]
        field_names = [f["name"] for f in fields]
        usage = LLMUsage()
        t0 = time.perf_counter()
        rows = await extract_from_document(parsed, fields, usage)
        elapsed = time.perf_counter() - t0
        report = print_extraction_report(rows, field_names, expected, f"Fields: {fs_name}", elapsed, usage.cost_usd)
        report["field_set"] = fs_name
        results.append(report)

    return results


async def main() -> None:
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--files", nargs="*", default=None, help="File keys to test (01-05). Default: all available")
    parser.add_argument("--skip-extraction", action="store_true", help="Skip full extraction, only test spreadsheet logic")
    parser.add_argument("--robustness", action="store_true", help="Run multi-field-set robustness tests")
    args = parser.parse_args()

    file_keys = args.files or [k for k in TEST_FILES if TEST_FILES[k][0].exists()]
    if not file_keys:
        print(f"{RED}No test files available{RESET}")
        return

    print(f"\n{BOLD}{'█' * 72}{RESET}")
    print(f"{BOLD}  SOV End-to-End Test Suite{RESET}")
    print(f"  Files: {file_keys}")
    print(f"  Skip extraction: {args.skip_extraction}")
    print(f"{BOLD}{'█' * 72}{RESET}")

    all_results: Dict[str, Any] = {}

    if not args.skip_extraction:
        for key in file_keys:
            fs_name = RECOMMENDED_FIELD_SET_PER_FILE.get(key, "frontend_defaults")
            report1 = await test_create_from_scratch(key, fs_name)
            all_results[f"create_{key}_{fs_name}"] = report1

            if report1 and report1.get("rows", 0) > 2:
                path, expected = TEST_FILES[key]
                fields = FIELD_SETS[fs_name]
                field_names = [f["name"] for f in fields]
                parsed = parse_pdf(str(path), path.name)
                usage = LLMUsage()
                rows = await extract_from_document(parsed, fields, usage)
                real_rows = [r for r in rows if not r.get("_error")]
                report2 = await test_baseline_update(key, fs_name, real_rows)
                all_results[f"update_{key}_{fs_name}"] = report2

        if args.robustness:
            for key in file_keys:
                robustness = await test_multi_field_set_extraction(key)
                all_results[f"robustness_{key}"] = robustness
    else:
        print(f"\n{YELLOW}  Skipping extraction, testing spreadsheet logic only{RESET}")
        for fs_name in ["frontend_defaults", "abbreviated", "verbose", "mixed_case_symbols"]:
            fields = FIELD_SETS[fs_name]
            field_names = [f["name"] for f in fields]
            fake_rows = []
            for i in range(5):
                row = {"_source_file": f"test_{i}.pdf"}
                for fn in field_names:
                    row[fn] = f"val_{fn}_{i}"
                fake_rows.append(row)

            print(f"\n  {CYAN}Testing spreadsheet ops with field set: {fs_name}{RESET}")

            excel_bytes = generate_excel_bytes(fake_rows, field_names)
            wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
            ws = wb.active
            data_rows = ws.max_row - 1
            print(f"    Excel: {data_rows} rows, headers OK: {ws.cell(row=1, column=2).value == field_names[0]}")

            csv_bytes = generate_csv_bytes(fake_rows, field_names)
            reader = csv.DictReader(io.StringIO(csv_bytes.decode("utf-8")))
            csv_rows = list(reader)
            print(f"    CSV: {len(csv_rows)} rows")

            new_rows = [{"_source_file": "update.pdf"} | {fn: f"new_{fn}" for fn in field_names}]
            updated = update_excel_baseline_bytes(excel_bytes, new_rows, field_names, True)
            wb2 = openpyxl.load_workbook(io.BytesIO(updated))
            print(f"    Excel update (edit=True): {wb2.active.max_row - 1} rows")

            preserved = update_excel_baseline_bytes(excel_bytes, new_rows, field_names, False)
            wb3 = openpyxl.load_workbook(io.BytesIO(preserved))
            print(f"    Excel update (edit=False): {wb3.active.max_row - 1} rows")

            csv_updated = update_csv_baseline_bytes(csv_bytes, new_rows, field_names, True)
            r2 = list(csv.DictReader(io.StringIO(csv_updated.decode("utf-8"))))
            print(f"    CSV update (edit=True): {len(r2)} rows")

            csv_preserved = update_csv_baseline_bytes(csv_bytes, new_rows, field_names, False)
            r3 = list(csv.DictReader(io.StringIO(csv_preserved.decode("utf-8"))))
            print(f"    CSV update (edit=False): {len(r3)} rows")

    # ── Final summary ──
    print(f"\n\n{BOLD}{'█' * 72}{RESET}")
    print(f"{BOLD}  FINAL SUMMARY{RESET}")
    print(f"{'█' * 72}{RESET}")

    pass_count = 0
    fail_count = 0
    for name, result in all_results.items():
        if result is None:
            continue
        if isinstance(result, dict):
            if result.get("pass"):
                print(f"  {GREEN}✓ {name}{RESET}")
                pass_count += 1
            elif result.get("count_ok") or result.get("fill_ok"):
                print(f"  {YELLOW}~ {name} (partial){RESET}")
                fail_count += 1
            elif "excel_edit_ok" in result:
                all_ok = result.get("excel_edit_ok") and result.get("csv_edit_ok")
                if all_ok:
                    print(f"  {GREEN}✓ {name}{RESET}")
                    pass_count += 1
                else:
                    print(f"  {YELLOW}~ {name}{RESET}")
                    fail_count += 1
            else:
                print(f"  {RED}✗ {name}{RESET}")
                fail_count += 1
        elif isinstance(result, list):
            for r in result:
                fs = r.get("field_set", "?")
                if r.get("pass"):
                    print(f"  {GREEN}✓ {name}/{fs}{RESET}")
                    pass_count += 1
                else:
                    print(f"  {RED}✗ {name}/{fs}{RESET}")
                    fail_count += 1

    total = pass_count + fail_count
    print(f"\n  {BOLD}Results: {pass_count}/{total} passed{RESET}")
    if fail_count == 0:
        print(f"  {GREEN}All tests passed!{RESET}")
    else:
        print(f"  {YELLOW}{fail_count} tests need attention{RESET}")

    print(f"\n{BOLD}Done.{RESET}\n")


if __name__ == "__main__":
    asyncio.run(main())
